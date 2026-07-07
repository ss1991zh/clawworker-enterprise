# -*- coding: utf-8 -*-
"""
公式列求值 —— 摄取阶段兜底。

问题背景
--------
用户上传的 Excel 常把"派生指标"(销售收入 = 销量×单价、营业利润率 = 营业利润/销售收入 …)
写成**公式列**。若这张表是用工具生成 / LibreOffice 保存 / 未在 Excel 里按过一次计算,
公式单元格就**没有缓存值**:openpyxl(以及 pandas.read_excel)读回来全是 None/NaN。

结果:这些派生列摄取时被判为"空字符串列"丢进 metadata,加密库里只有原始录入列。
下游 AI 生成的代码 `df["营业利润率"]` 读到的是空列 → 计算结果全 NaN → 明文为空、
密文是 encrypt(0)。同一张表若在 Excel 里存过一次(公式已缓存值),就一切正常
("以前是好的"就是这个区别)。

本模块的职责
------------
摄取时,对**整列为空的公式列**,用同行引用的其它列**把公式算出来**,填回 DataFrame,
使其变回真实数值列 → 正常加密 → 下游计算正确。

设计要点
--------
- 只处理 xlsx;只填**当前整列为 NaN** 的公式列(有缓存值的列一律不动,避免改变正确数据)。
- Excel 列字母 ↔ df 列按**位置**对应(A=第1列),并用表头名做一致性校验,不一致就整体放弃。
- 公式 → 受限 Python 表达式(pandas 向量化):支持 + - * / ( )、比较、
  IF/SUM/ABS/ROUND/MIN/MAX/MAX0。用 ast 白名单校验,杜绝任意代码执行。
- 依赖顺序自动解析(H=F*G,K=H-J,O=N/H … 多趟迭代直到收敛)。
- 任一步失败都安全兜底:该列保持原样(仍为空,后续照旧丢 metadata),绝不阻断摄取。
"""

from __future__ import annotations

import ast
import re
from typing import Optional


# Excel 函数 → 内部实现名
_FUNCS = {
    "IF": "_IF", "SUM": "_SUM", "ABS": "_ABS",
    "ROUND": "_ROUND", "MIN": "_MIN", "MAX": "_MAX",
}

# ast 白名单:表达式里只允许这些节点
_ALLOWED_NODES = (
    ast.Expression, ast.BinOp, ast.UnaryOp, ast.BoolOp, ast.Compare,
    ast.Call, ast.Name, ast.Load, ast.Subscript, ast.Constant,
    ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow,
    ast.USub, ast.UAdd, ast.And, ast.Or,
    ast.Eq, ast.NotEq, ast.Lt, ast.Gt, ast.LtE, ast.GtE,
)
_ALLOWED_NAMES = {"V"} | set(_FUNCS.values())


def _cell_letter(col_index_zero_based: int) -> str:
    """0 基列号 → Excel 列字母(0→A,25→Z,26→AA)。"""
    n = col_index_zero_based + 1
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _translate(formula: str, letters_used: set) -> str:
    """Excel 公式字符串 → 受限 Python 表达式;把用到的列字母塞进 letters_used。"""
    s = formula.strip()
    if s.startswith("="):
        s = s[1:]
    s = s.replace("$", "")
    # 比较符归一:先护住 <= >= <>,再把单个 = 变 ==
    s = s.replace("<=", "\x01").replace(">=", "\x02").replace("<>", "!=")
    s = s.replace("=", "==")
    s = s.replace("\x01", "<=").replace("\x02", ">=")

    # 单元格引用 A1 / AB12 → V["A"](同行引用,只取列字母)。
    # 函数名(IF/SUM…)后面跟 '(' 不跟数字,不会被这条命中。
    def _repl_ref(m: "re.Match") -> str:
        col = m.group(1).upper()
        letters_used.add(col)
        return f'V["{col}"]'

    s = re.sub(r"\b([A-Za-z]{1,3})\d+\b", _repl_ref, s)

    # 函数名映射(大小写不敏感)
    for xl, py in _FUNCS.items():
        s = re.sub(rf"\b{xl}\s*\(", f"{py}(", s, flags=re.IGNORECASE)
    return s


def _check_safe(expr: str) -> bool:
    """ast 白名单校验:只允许算术 / 比较 / 受限函数调用 / V[...] 取列。"""
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError:
        return False
    for node in ast.walk(tree):
        if not isinstance(node, _ALLOWED_NODES):
            return False
        if isinstance(node, ast.Name) and node.id not in _ALLOWED_NAMES:
            return False
        if isinstance(node, ast.Call):
            # 只允许调用白名单函数名(不允许 V(...) 之类)
            if not (isinstance(node.func, ast.Name) and node.func.id in set(_FUNCS.values())):
                return False
        if isinstance(node, ast.Subscript):
            # 只允许 V["字面量字符串"]
            if not (isinstance(node.value, ast.Name) and node.value.id == "V"):
                return False
    return True


def fill_formula_columns(
    path: str,
    df,
    *,
    sheet_name: Optional[str] = None,
    header_row: int = 0,
    log=None,
) -> tuple:
    """
    对 df 中"整列为空的公式列"求值填回。返回 (df, filled_columns)。

    path        : 源 xlsx 路径(用 openpyxl 读公式字符串)
    df          : pandas.read_excel 读回的 DataFrame(公式列此时是 NaN)
    sheet_name  : _smart_read 选中的 sheet 名
    header_row  : 表头所在行(0 基,pandas 语义)
    filled_columns : 实际算出并填回的列名列表(空列表表示没动)
    """
    filled: list = []
    try:
        import numpy as np
        import pandas as pd
        import openpyxl
    except Exception:
        return df, filled

    # 只处理 xlsx
    if not str(path).lower().endswith((".xlsx", ".xlsm", ".xls")):
        return df, filled

    # 只有存在整列全空的列才值得费劲(否则直接返回)
    empty_cols = [c for c in df.columns if df[c].isna().all()]
    if not empty_cols:
        return df, filled

    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return df, filled
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

        # 列字母 ↔ df 列名(按位置),并用表头名做一致性校验
        hdr_excel_row = header_row + 1          # openpyxl 1 基
        data_excel_row = header_row + 2         # 首个数据行
        col_letter_to_name: dict = {}
        header_cells = {c.column_letter: c.value for c in ws[hdr_excel_row]}
        for i, name in enumerate(df.columns):
            letter = _cell_letter(i)
            hv = header_cells.get(letter)
            # 表头名对不上(位置漂移 / 合并单元格)→ 整体放弃,安全兜底
            if hv is None or str(hv).strip() != str(name).strip():
                return df, filled
            col_letter_to_name[letter] = name
        name_to_letter = {v: k for k, v in col_letter_to_name.items()}

        # 读首个数据行,收集"整列为空且该行是公式"的列 → 待求值
        first_row_cells = {c.column_letter: c for c in ws[data_excel_row]}
        to_fill: dict = {}     # letter -> (colname, formula_str)
        for c in empty_cols:
            letter = name_to_letter.get(c)
            if not letter:
                continue
            cell = first_row_cells.get(letter)
            if cell is None:
                continue
            val = cell.value
            is_formula = (getattr(cell, "data_type", None) == "f") or \
                         (isinstance(val, str) and val.startswith("="))
            if is_formula and isinstance(val, str):
                to_fill[letter] = (c, val)
        if not to_fill:
            return df, filled

        # 可用列字母集合:一开始 = 所有"非待求值"列(原始录入列 + 身份列)
        available = {lt for lt in col_letter_to_name if lt not in to_fill}

        # V:列字母 → 数值 Series(按需构建)
        def _series(letter: str):
            return pd.to_numeric(df[col_letter_to_name[letter]], errors="coerce")

        _IF = lambda cond, a, b: np.where(cond, a, b)              # noqa: E731
        _SUM = lambda *a: np.add.reduce([np.asarray(x, float) for x in a])  # noqa: E731
        _ABS = np.abs
        _ROUND = lambda x, n=0: np.round(x, int(n))                # noqa: E731
        _MIN = lambda *a: np.minimum.reduce([np.asarray(x, float) for x in a])  # noqa: E731
        _MAX = lambda *a: np.maximum.reduce([np.asarray(x, float) for x in a])  # noqa: E731

        # 多趟迭代解析依赖(H=F*G,K=H-J,O=N/H …)
        for _ in range(len(to_fill) + 2):
            if not to_fill:
                break
            progressed = False
            for letter in list(to_fill.keys()):
                colname, formula = to_fill[letter]
                letters_used: set = set()
                try:
                    expr = _translate(formula, letters_used)
                except Exception:
                    del to_fill[letter]         # 无法翻译 → 放弃该列
                    continue
                if not letters_used or not letters_used.issubset(available):
                    continue                    # 依赖尚未就绪,下一趟再试
                if not _check_safe(expr):
                    del to_fill[letter]
                    continue
                V = {lt: _series(lt) for lt in letters_used}
                ns = {
                    "V": V, "_IF": _IF, "_SUM": _SUM, "_ABS": _ABS,
                    "_ROUND": _ROUND, "_MIN": _MIN, "_MAX": _MAX,
                }
                try:
                    res = eval(expr, {"__builtins__": {}}, ns)  # noqa: S307 —— ast 已白名单校验
                    ser = pd.Series(np.asarray(res, dtype="float64"), index=df.index)
                    ser = ser.replace([np.inf, -np.inf], np.nan)
                except Exception:
                    del to_fill[letter]
                    continue
                if ser.notna().any():
                    df[colname] = ser
                    filled.append(colname)
                    available.add(letter)
                del to_fill[letter]
                progressed = True
            if not progressed:
                break
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if filled and log:
        try:
            log("think", "已按源表公式补算派生列:" + "、".join(filled))
        except Exception:
            pass
    return df, filled
