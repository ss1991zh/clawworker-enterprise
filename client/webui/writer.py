"""
v2 Excel writer — 把 [(sheet_name, DataFrame, chart_hint)] 直接写成多 sheet xlsx。

不依赖 LangGraph / workflow / ops。每个 skill 函数产出的 (name, df, chart)
被一一渲染:
- 数字列自动格式(率→0.00% / 金额→#,##0.00 / 天数→0.0 / 其他→0)
- 如果 chart_hint 含 type/x/y 就加柱状图或折线图
- 输出落到 ~/Downloads/analysis_<ts>.xlsx
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional


# 路径白名单(B6 §2)
# 交互式输出暂存目录:Excel 先写这里(不自动落 Downloads),用户点「下载」才存到 Downloads
OUTPUTS_DIR = Path.home() / ".agent-system" / "outputs"

_ENC_ROW_LIMIT = 50_000   # 密文版 Excel 单表行数上限(再加密慢+文件巨大),超则截断并标注


def make_excel_path(stem: Optional[str] = None, staging: bool = False) -> Path:
    """
    生成 Excel 落盘路径。
    staging=False(默认):~/Downloads/<stem>_<ts>.xlsx(批量解密到文件夹等显式产出)。
    staging=True:~/.agent-system/outputs/<stem>_<ts>.xlsx(交互式生成,不自动进 Downloads,
                 用户点「下载」时再由浏览器存到 Downloads)。
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = (stem or "analysis").strip("_") or "analysis"
    base = OUTPUTS_DIR if staging else (Path.home() / "Downloads")
    return base / f"{safe_stem}_{ts}.xlsx"


# ----------------------------------------------------------------------------
# 文件名衍生:<密文文件名> + <用户需求关键词> + <时间戳>
# ----------------------------------------------------------------------------

# 用户问题里要去掉的"动词/语气词前缀",防止文件名一上来就是「计算...」
_QUERY_VERBS = (
    "计算", "统计", "分析", "汇总", "排名", "排行", "列出", "生成", "导出",
    "看一下", "看下", "看看", "看", "算一下", "算下", "算出", "算", "求一下", "求",
    "查询", "查一下", "查", "做一下", "做下", "做",
    "帮我", "帮忙", "请", "我要", "我想", "我要看", "给我", "麻烦",
)
# 文件名非法字符 / 影响阅读的标点
_BAD_FNAME_CHARS = re.compile(r"[\\/:*?\"<>|\s,.，。;；·!?？“”‘’\[\]\(\)（）【】]+")


def _clean_cipher_stem(cipher_path: Path) -> str:
    """从密文文件路径里抠出"原文件名":去掉 _enc 后缀 / .cipher / 各种扩展。"""
    stem = cipher_path.stem  # 销售提成模拟数据_管理会计_enc
    # 去掉常见加密尾缀(_enc / _encrypted / .cipher 子串)
    for suffix in ("_enc", "_encrypted", ".cipher", "_encrypt"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
    # 再次去 stem 上的后缀(.xlsx.cipher 这种)
    while Path(stem).suffix:
        stem = Path(stem).stem
    return stem.strip("_") or "data"


def _clean_query_keyword(user_query: str, max_len: int = 14) -> str:
    """从用户问题里抽一段短关键词作为文件名一部分。"""
    q = (user_query or "").strip()
    if not q:
        return ""
    # 去掉前缀动词
    for v in _QUERY_VERBS:
        if q.startswith(v):
            q = q[len(v):].strip()
            break
    # 去标点 / 空白
    q = _BAD_FNAME_CHARS.sub("", q)
    # 截断到 max_len 字符
    if len(q) > max_len:
        q = q[:max_len]
    return q.strip("_")


def derive_excel_stem(cipher_path: Optional[Path], user_query: str) -> str:
    """
    返回 Excel 文件名的 stem(无 _时间戳 部分)。
    形如:销售提成模拟数据_管理会计_销售提成
    """
    parts: list[str] = []
    if cipher_path is not None:
        cipher_stem = _clean_cipher_stem(cipher_path)
        if cipher_stem:
            parts.append(cipher_stem)
    kw = _clean_query_keyword(user_query)
    if kw:
        parts.append(kw)
    if not parts:
        return "analysis"
    stem = "_".join(parts)
    # 文件系统兜底:不要太长(Windows / macOS 都有 255 字节文件名上限)
    if len(stem) > 80:
        stem = stem[:80].rstrip("_")
    return stem


# 用户为定时任务显式指定的输出文件夹根(由 app 在 create/patch/startup 注册)。
# 这是用户主动选择的落盘位置(等同选择下载目录),纳入白名单。
_EXTRA_OUTPUT_ROOTS: set[Path] = set()


def register_output_root(folder) -> None:
    """登记一个用户指定的输出文件夹根,使其下的 Excel 写入通过白名单校验。"""
    try:
        if folder:
            _EXTRA_OUTPUT_ROOTS.add(Path(folder).expanduser().resolve())
    except Exception:
        pass


def enforce_excel_path(path: Path) -> Path:
    """B6 §2:Excel 输出强制在白名单目录 —— ~/Downloads/、暂存 ~/.agent-system/outputs/,
    以及用户为定时任务显式指定并已登记的输出文件夹。"""
    resolved = path.expanduser().resolve()
    roots = [(Path.home() / "Downloads").resolve(), OUTPUTS_DIR.resolve(), *_EXTRA_OUTPUT_ROOTS]
    for r in roots:
        try:
            resolved.relative_to(r)
            return resolved
        except ValueError:
            continue
    raise PermissionError(f"Excel 输出路径不在白名单内:{resolved}")


# ----------------------------------------------------------------------------
# 产品级呈现:number_format 推断 + 语义调色板 + 列宽
# ----------------------------------------------------------------------------

# 逆向指标:数值越大越"坏",百分比色阶要反色(红高绿低)
_REVERSE_METRIC = ("逾期", "超支", "流失", "异常", "缺货", "呆滞", "波动", "降幅")
# 方向取决于是成本类还是收入类,通用渲染层无从判断(收入差异率高=好、成本差异率高=坏)——
# 这类列不套三色阶,红绿由方向准确的档位列(超额/超支等)承担,避免把"收入超额"误染成红。
_AMBIGUOUS_DIRECTION = ("差异率", "差额率", "偏差率", "变动率")


# 金额格式:正数千分位,负数会计惯例红字括号
_MONEY_FMT = "#,##0.00;[Red](#,##0.00)"


def _rotated_text_props(deg: int):
    """构造图表轴标签旋转的 txPr(角度单位 1/60000 度)。"""
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (
        CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties,
    )
    return RichText(
        bodyPr=RichTextProperties(rot=int(deg * 60000), vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties()), endParaRPr=None)],
    )


def _infer_number_format(col_name: str) -> Optional[str]:
    """按列名推断 Excel 数字格式 —— 覆盖销售/财务/库存/HR/客户常见指标。"""
    name = str(col_name)
    lower = name.lower()
    # 百分比 —— 注意:「同比增长/环比增长」是**绝对增长额(元)**不是比率,
    # 只有带「率」或明确「涨幅/降幅/占比」的才是百分比(下面金额段会接住增长额)
    if any(k in lower for k in ("rate", "ratio", "percentage", "pct")):
        return "0.00%"
    if any(k in name for k in ("率", "比例", "占比", "百分比", "涨幅", "降幅")):
        return "0.00%"
    # 天数 / 账龄 / 周转 —— 但「账龄未知」这类**账龄分桶列装的是金额**,不是天数,别误判
    if not any(k in name for k in ("未知", "未到期")):
        if "days" in lower or any(k in name for k in ("天数", "账龄", "周转")):
            return "0.0"
    # 金额(财务/销售/HR 薪酬)—— 负数用会计惯例红字括号 (1,234.00)
    if any(k in lower for k in ("amount", "value", "price", "cost", "revenue", "profit", "sales", "budget")):
        return _MONEY_FMT
    if any(k in name for k in (
        "金额", "额", "收入", "成本", "利润", "营收", "回款", "应收", "应付", "余额",
        "预算", "实际", "增长", "增减", "差额", "净增", "净额",   # 同比/环比增长额等
        "(元)", "（元）", "工资", "薪酬", "薪资", "提成", "奖金", "单价", "总价",
    )):
        return _MONEY_FMT
    # 计数 / 人数 / 排名(整数)
    if any(k in lower for k in ("count", "num", "qty")):
        return "0"
    if any(k in name for k in ("数量", "笔数", "次数", "订单数", "人数", "件数", "排名", "名次", "人次", "户数")):
        return "0"
    return None


# 语义档位 → (背景色, 字体色)。覆盖销售/财务/库存/HR/客户各场景的常见标签。
_TIER_BEST = ("00B050", "FFFFFF")   # 深绿:头档(优/优秀)—— 与"良"的浅绿区分,四档不塌成三色
_TIER_GOOD = ("C6EFCE", "006100")   # 绿:达成 / 正常 / 良 / 健康 / 活跃 / A / 节约 / 盈利
_TIER_WARN = ("FFEB9C", "9C5700")   # 黄:预警 / 关注 / 中 / B / 临期 / 潜力
_TIER_BAD  = ("FFC7CE", "9C0006")   # 红:未达成 / 异常 / 超支 / 逾期 / 呆滞 / 流失 / C / 亏损
_TIER_INFO = ("DDEBF7", "1F4E78")   # 蓝:预测 / 新增 / 重点(中性强调)

_TIER_MAP = {
    # —— 深绿(头档)——
    "优": _TIER_BEST, "优秀": _TIER_BEST,
    # —— 绿(好)——
    "达成": _TIER_GOOD, "已达成": _TIER_GOOD, "完成": _TIER_GOOD, "正常": _TIER_GOOD,
    "健康": _TIER_GOOD, "活跃": _TIER_GOOD, "超额": _TIER_GOOD, "达标": _TIER_GOOD,
    "盈利": _TIER_GOOD, "节约": _TIER_GOOD, "favorable": _TIER_GOOD, "a": _TIER_GOOD,
    "a类": _TIER_GOOD, "a档": _TIER_GOOD, "重要价值": _TIER_GOOD, "高价值": _TIER_GOOD,
    "高": _TIER_GOOD, "充足": _TIER_GOOD, "良": _TIER_GOOD,
    # —— 黄(关注)——
    "预警": _TIER_WARN, "关注": _TIER_WARN, "中": _TIER_WARN, "一般": _TIER_WARN,
    "中等": _TIER_WARN, "临期": _TIER_WARN, "b": _TIER_WARN, "b类": _TIER_WARN,
    "b档": _TIER_WARN, "潜力": _TIER_WARN, "待挽留": _TIER_WARN, "一般价值": _TIER_WARN,
    # —— 红(差/风险)——
    "未达成": _TIER_BAD, "未完成": _TIER_BAD, "未达": _TIER_BAD, "异常": _TIER_BAD, "超支": _TIER_BAD,
    "逾期": _TIER_BAD, "呆滞": _TIER_BAD, "流失": _TIER_BAD, "已流失": _TIER_BAD,
    "差": _TIER_BAD, "亏损": _TIER_BAD, "高风险": _TIER_BAD, "unfavorable": _TIER_BAD,
    "c": _TIER_BAD, "c类": _TIER_BAD, "c档": _TIER_BAD, "低": _TIER_BAD, "缺货": _TIER_BAD,
    # —— 蓝(中性强调)——
    "预测": _TIER_INFO, "新增": _TIER_INFO, "新客": _TIER_INFO, "重点": _TIER_INFO,
    "历史": None,  # 历史不上色
}

# 自动识别"档位文字列"的列名关键词(各场景通用)
_TIER_COL_HINTS = ("类型", "档位", "等级", "分类", "分群", "状态", "评级", "是否", "标签", "情况", "abc")


# 否定前缀:出现在档位词之前会翻转语义("不活跃"=差、"未流失"=好)
_TIER_NEGATIONS = ("不", "未", "非", "无", "没")


def _tier_fill(value):
    """
    档位文字 → (PatternFill, Font) 或 None。
    两个防误染:
      ① 否定词感知:匹配到的档位词前若紧跟 不/未/非/无/没,好↔差**翻转**
         (否则"不活跃"含"活跃"被染绿、"未流失"含"流失"被染红,与真实语义相反);
      ② 单字键(高/中/低/优/良/差/a/b/c)只做**精确**匹配,不参与模糊包含
         (否则"华中"含"中"、"Stable"含"a" 被乱染)。
    """
    from openpyxl.styles import Font, PatternFill
    if value is None:
        return None
    key = str(value).strip().lower()
    if not key:
        return None
    pal = _TIER_MAP.get(key)   # 精确匹配(单字键也在此命中,如整格就是"高")
    if pal is None:
        # 模糊包含:仅多字键参与;命中后按前缀否定翻转
        for k, p in _TIER_MAP.items():
            if p is None or len(k) <= 1:
                continue
            idx = key.find(k)
            if idx < 0:
                continue
            # 否定检测:向档位词左侧回看一个小窗口(而非仅紧邻前一字),
            # 覆盖"未能达成/没有完成/不太活跃/并非达成"等隔字否定 → 好↔差翻转
            window = key[max(0, idx - 3):idx]
            negated = any(n in window for n in _TIER_NEGATIONS)
            pal = _flip_tier(p) if negated else p
            break
    if not pal:
        return None
    bg, fg = pal
    return PatternFill("solid", fgColor=bg), Font(color=fg, bold=True)


def _flip_tier(pal):
    """好↔差翻转;中性/关注档(黄/蓝)否定后语义模糊 → 不上色。"""
    if pal in (_TIER_GOOD, _TIER_BEST):
        return _TIER_BAD
    if pal == _TIER_BAD:
        return _TIER_GOOD
    return None


def _disp_width(s) -> float:
    """显示宽度:中文/全角算 1.8,其余算 1.0。"""
    w = 0.0
    for ch in str(s):
        w += 1.8 if ord(ch) > 0x2E7F else 1.0
    return w


# CSV/公式注入前缀:Excel 把以这些字符开头的字符串当公式执行(DDE/HYPERLINK 钓鱼/数据外带)
_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _defuse_formula(s: str) -> str:
    """身份列文本若以危险字符开头,前置单引号,让 Excel 当纯文本而非公式。"""
    if s and s[0] in _INJECTION_PREFIXES:
        return "'" + s
    return s


def _clean_val(v):
    """numpy 标量 → python;NaN/NaT → None(避免 Excel 里出现 'nan');字符串防公式注入。"""
    try:
        import pandas as pd
        if v is None:
            return None
        if not isinstance(v, (str, bytes, list, tuple, dict)) and pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, str):
        return _defuse_formula(v)      # 解密后的身份列文本可能夹带 =cmd|... 之类注入载荷
    if hasattr(v, "item") and not isinstance(v, (str, bytes)):
        try:
            return v.item()
        except Exception:
            return v
    return v


# ----------------------------------------------------------------------------
# 主入口
# ----------------------------------------------------------------------------

def write_skill_results(
    results: list[dict],
    path: Optional[Path] = None,
    stem: Optional[str] = None,
    staging: bool = False,
    provenance: Optional[dict] = None,
) -> Path:
    """
    把 [{sheet_name, df, ...}] 列表写成产品级多 sheet xlsx。
    staging=True:写到暂存目录(不自动进 Downloads,供交互式「点下载才保存」)。

    每个 result dict 至少含 sheet_name + df,以下键全部可选(声明即美化):
        chart:          {"type":"bar"|"line"|"pie"|"stacked_bar","x":..,"y":..|[..],"title":..}
        tier_col:       <档位文字列名> —— 该列按语义上色(不传则按列名自动识别)
        total_row:      True | {"label":"合计","sum":[列..],"mean":[列..]}
        note:           <表顶一行灰色说明>
        number_formats: {"列名":"0.00%"} —— 显式覆盖某列格式

    渲染端统一施加:表头高亮 / 冻结首行 / 自动筛选 / 隔行底色 /
    内容感知列宽 / 百分比列三色阶(逆向指标反色)/ 图表增强。
    """
    import openpyxl

    dst = enforce_excel_path(path or make_excel_path(stem, staging=staging))
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删默认 sheet

    for r in results:
        df = r.get("df")
        if df is None or getattr(df, "empty", True):
            continue
        df = _dedup_columns(df)  # 防御:同名列(如误重复 merge 身份列)只留第一份
        sheet_name = str(r.get("sheet_name") or "Sheet")[:31]
        ws = wb.create_sheet(title=sheet_name)
        try:
            _render_sheet(ws, df, r)
        except Exception:
            # 装饰失败 → 退回最简写法,保证数据不丢、文件可落盘
            _render_plain(ws, df)

    if not wb.sheetnames:  # 全空兜底(openpyxl 不允许零 sheet 保存)
        wb.create_sheet(title="结果")

    # 溯源 sheet:这份 Excel 用什么数据、什么口径、什么时候算的 —— 审计可回答
    if provenance is not None:
        try:
            _append_provenance_sheet(wb, results, provenance)
        except Exception:
            pass  # 溯源失败不阻断落盘

    wb.save(dst)
    return dst


def _append_provenance_sheet(wb, results: list[dict], provenance: dict) -> None:
    """末尾附「口径说明」sheet:生成时间 + 调用方提供的溯源信息 + 各 sheet 行列数与口径。"""
    from datetime import datetime
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title="口径说明")
    rows: list[tuple[str, str]] = [("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]
    rows.extend((str(k), str(v)) for k, v in provenance.items())
    for r in results:
        df = r.get("df")
        if df is None or getattr(df, "empty", True):
            continue
        nm = str(r.get("sheet_name") or "Sheet")[:31]
        desc = f"{len(df)} 行 × {len(df.columns)} 列"
        if r.get("note"):
            desc += f" · {r['note']}"
        rows.append((f"sheet「{nm}」", desc))

    head = ws.cell(1, 1, "结果溯源 · 口径说明")
    head.font = Font(bold=True, size=13, color="1E3A8A")
    head.fill = PatternFill("solid", fgColor="DBEAFE")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    for i, (k, v) in enumerate(rows, 2):
        kc = ws.cell(i, 1, k)
        kc.font = Font(bold=True, color="374151")
        vc = ws.cell(i, 2, v)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90


def _dedup_columns(df):
    """同名列只留第一份(防 LLM 误把身份列 merge 两次导致重列)。"""
    try:
        if getattr(df, "columns", None) is not None and df.columns.duplicated().any():
            return df.loc[:, ~df.columns.duplicated()]
    except Exception:
        pass
    return df


def _render_plain(ws, df) -> None:
    """最简兜底:只写表头 + 数据,不做任何样式。"""
    for ci, h in enumerate([str(c) for c in df.columns], 1):
        ws.cell(1, ci, h)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, _clean_val(val))


def _render_sheet(ws, df, r: dict) -> None:
    """渲染单张产品级 sheet。"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    band_fill = PatternFill("solid", fgColor="F2F6FC")
    center = Alignment(horizontal="center", vertical="center")

    # 图表规格(chart + charts)+ split_by 预排序:同组行连续 → 可按组拆成多张图
    chart_specs = _collect_chart_specs(r)

    # 兜底自动配图:没有任何有效图表提示(LLM 偶尔漏配)时,自动挑
    # 文本主体列 × 指标列配一张;行多自动按低基数维度 split_by 拆图
    def _spec_valid(s):
        x = s.get("x")
        y = s.get("y")
        ys = [y] if isinstance(y, str) else list(y or [])
        return x in df.columns and any(c in df.columns for c in ys)
    if not any(_spec_valid(s) for s in chart_specs):
        auto = _auto_chart_spec(df)
        chart_specs = [auto] if auto else []

    split_col = next(
        (s.get("split_by") for s in chart_specs
         if s.get("split_by") and s.get("split_by") in df.columns),
        None,
    )
    if split_col:
        try:
            df = df.sort_values(split_col, kind="stable").reset_index(drop=True)
        except Exception:
            pass

    headers = [str(c) for c in df.columns]
    n_cols = len(headers)
    header_row = 1

    # 可选:表顶说明 caption(时间列有断档时自动追加缺期预警)
    note = r.get("note")
    gap_warn = _time_gap_warning(df, chart_specs)
    if gap_warn:
        note = f"{note} · {gap_warn}" if note else gap_warn
    if note:
        c = ws.cell(1, 1, str(note))
        c.font = Font(italic=True, color="6B7280", size=10)
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        header_row = 2

    # 表头
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(header_row, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[header_row].height = 20

    # 每列 number_format(+ 显式覆盖)
    col_fmt: dict[int, str] = {}
    for ci, h in enumerate(headers, 1):
        nf = _infer_number_format(h)
        if nf == "0.00%":
            # 百分比口径体检:列名像"率"但数据存成整数百分数(85 表示 85%)而非小数(0.85)。
            # 判据要**保守**:超额完成率(1.6=160%)是合法小数,不能误判成整数百分数缩小 100 倍。
            # 只有当中位数明显大(>5)且几乎无小于 1.5 的值(整数百分数列不会有 0.x)时才认定。
            # 单行结果表(汇总/单实体)也要体检 —— 否则 85 会被当小数渲染成 8500%;
            # 中位数>5 的门槛已足够挡住合法小数(1.6=160% 不会误判)。
            try:
                import pandas as pd
                v = pd.to_numeric(df[h], errors="coerce").abs()
                v = v[v.notna()]
                # 零值对"整数百分数 vs 小数"没有判别力(两种口径下都是 0),必须排除:
                # 否则 [85,92,88,75,0] 里那个合法的 0 把占比拉到 0.8 < 0.9,体检失效,
                # 85 被当小数渲染成 8500%。
                nz = v[v > 0]
                if len(nz) >= 1 and float(nz.median()) > 5 and float((nz > 1.5).mean()) >= 0.9:
                    nf = '0.00"%"'
            except Exception:
                pass
        # 列名推断不出格式,但该列是**大数值**列(如账龄桶金额、透视里以产品线命名的金额列)
        # → 兜底给千分位,避免 1234567890 显示成科学计数 1.23E+09
        if nf is None:
            try:
                import pandas as pd
                s = pd.to_numeric(df[h], errors="coerce")
                s = s[s.notna()]
                if len(s) and float(s.abs().max()) >= 10000:
                    nf = _MONEY_FMT if (s != s.round()).any() else "#,##0"
            except Exception:
                pass
        if nf:
            col_fmt[ci] = nf
    for col, fmt in (r.get("number_formats") or {}).items():
        if col in headers:
            col_fmt[headers.index(col) + 1] = fmt

    # 档位列(显式 tier_col 优先,否则按列名自动识别)
    tier_col = r.get("tier_col")
    if not tier_col:
        for h in headers:
            if any(k in str(h).lower() for k in _TIER_COL_HINTS):
                tier_col = h
                break
    tier_ci = headers.index(tier_col) + 1 if tier_col in headers else None

    # 数据行
    data0 = header_row + 1
    for ri, row in enumerate(df.itertuples(index=False), data0):
        banded = (ri - data0) % 2 == 1
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, _clean_val(val))
            if ci in col_fmt:
                cell.number_format = col_fmt[ci]
            if banded:
                cell.fill = band_fill
            if tier_ci and ci == tier_ci:
                tf = _tier_fill(val)
                if tf:
                    cell.fill, cell.font = tf
    n_data = len(df)
    data_last = data0 + n_data - 1

    # 可选合计行
    if r.get("total_row"):
        try:
            _write_total_row(ws, df, headers, col_fmt, data0, data_last, r.get("total_row"))
        except Exception:
            pass

    # 内容感知列宽
    for ci, h in enumerate(headers, 1):
        w = _disp_width(h)
        try:
            for v in df.iloc[:, ci - 1].astype(str).head(60):
                w = max(w, _disp_width(v))
        except Exception:
            pass
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2.5, 10), 40)

    # 冻结表头(+ 宽表冻结身份首列:列多时,横向滚动仍见行标识)
    # 身份列判定:非数值列,或列名像身份(编号/工号/ID/姓名/物料/客户…)—— 后者覆盖
    # "员工编号被读成 int" 这类数值型身份列(否则会随滚动移出视野)
    freeze_col = 1
    try:
        import pandas as pd
        if n_cols > 6 and len(df.columns):
            first = str(df.columns[0])
            id_name = any(k in first for k in (
                "编号", "工号", "编码", "代码", "ID", "id", "姓名", "名称", "物料",
                "客户", "代表", "产品", "员工", "SKU", "单号", "订单号"))
            if id_name or not pd.api.types.is_numeric_dtype(df.iloc[:, 0]):
                freeze_col = 2
    except Exception:
        pass
    ws.freeze_panes = ws.cell(data0, freeze_col).coordinate
    if n_cols and n_data:
        ws.auto_filter.ref = (
            f"{ws.cell(header_row, 1).coordinate}:{get_column_letter(n_cols)}{data_last}"
        )

    # 百分比列三色阶(逆向指标如逾期/超支自动反色)
    if n_data >= 2:
        for ci, h in enumerate(headers, 1):
            if col_fmt.get(ci) not in ("0.00%", '0.00"%"'):
                continue
            if any(k in str(h) for k in _AMBIGUOUS_DIRECTION):
                continue   # 方向歧义列不套色阶,红绿交给档位列
            lo, mid, hi = "F8696B", "FFEB84", "63BE7B"  # 红→黄→绿
            if any(k in str(h) for k in _REVERSE_METRIC):
                lo, hi = hi, lo
            # 含负值(亏损)的列:色阶中点锚在 0,负=红正=绿;否则按中位数
            has_neg = False
            try:
                import pandas as pd
                if h in df.columns:
                    has_neg = bool((pd.to_numeric(df[h], errors="coerce") < 0).any())
            except Exception:
                pass
            col = get_column_letter(ci)
            try:
                rule = (ColorScaleRule(start_type="min", start_color=lo,
                                       mid_type="num", mid_value=0, mid_color=mid,
                                       end_type="max", end_color=hi)
                        if has_neg else
                        ColorScaleRule(start_type="min", start_color=lo,
                                       mid_type="percentile", mid_value=50, mid_color=mid,
                                       end_type="max", end_color=hi))
                ws.conditional_formatting.add(f"{col}{data0}:{col}{data_last}", rule)
            except Exception:
                pass

    # 图表(单图 / 多图 / 按 split_by 自动拆成每组一张,竖直堆叠在数据右侧)
    if chart_specs and n_data:
        try:
            _render_charts(ws, df, chart_specs, headers, header_row, n_cols)
        except Exception:
            pass  # 图表失败不阻断落盘


def _weighted_ratio_total(df, ratio_col):
    """
    比率列的合计口径:在表内找 分子/分母 列对(a/b 逐行 ≈ 比率列)→ 返回 sum(a)/sum(b)。
    合计行的总回款率 = 总回款 ÷ 总销售额(加权),**不是**各行率的简单平均
    (大客户 90% 小客户 10%,简单平均 50%,加权可能 85%)。匹配不到返回 None。
    """
    import numpy as np
    import pandas as pd
    try:
        r = pd.to_numeric(df[ratio_col], errors="coerce").to_numpy(dtype=float)
        mask = np.isfinite(r)
        if mask.sum() < 2:
            return None
        num_cols = [c for c in df.columns
                    if c != ratio_col and pd.api.types.is_numeric_dtype(df[c])]
        for a in num_cols:
            va = pd.to_numeric(df[a], errors="coerce").to_numpy(dtype=float)
            for b in num_cols:
                if a == b:
                    continue
                vb = pd.to_numeric(df[b], errors="coerce").to_numpy(dtype=float)
                with np.errstate(divide="ignore", invalid="ignore"):
                    q = va / vb
                ok = np.isfinite(q) & mask
                if ok.sum() < 2 or ok.sum() < mask.sum() * 0.9:
                    continue
                if np.allclose(q[ok], r[ok], rtol=5e-3, atol=1e-9):
                    sb = float(vb[ok].sum())
                    if sb:
                        return float(va[ok].sum()) / sb
        return None
    except Exception:
        return None


def _write_total_row(ws, df, headers, col_fmt, data0, data_last, spec) -> None:
    """
    在数据末尾追加一行合计/均值。spec=True 自动推断;dict 可指定
    {label, sum:[列..], mean:[列..], values:{列:显式值}}(values 优先级最高)。
    比率列自动尝试**加权口径**(总分子/总分母),匹配不到才退回简单均值。
    """
    import pandas as pd
    from openpyxl.styles import Font, PatternFill

    label = "合计"
    sum_cols: list = []
    mean_cols: list = []
    values: dict = {}
    if isinstance(spec, dict):
        label = spec.get("label") or "合计"
        sum_cols = list(spec.get("sum") or [])
        mean_cols = list(spec.get("mean") or [])
        values = dict(spec.get("values") or {})
    if not sum_cols and not mean_cols:  # 自动:金额/计数列求和,百分比列取均值
        for h in headers:
            # 排名/序号是序数,求和无意义,跳过
            if any(k in str(h) for k in ("排名", "名次", "序号")):
                continue
            fmt = _infer_number_format(h)
            if fmt in (_MONEY_FMT, "0"):
                sum_cols.append(h)
            elif fmt == "0.00%":
                mean_cols.append(h)

    row_idx = data_last + 1
    tot_font = Font(bold=True, color="1F2937")
    tot_fill = PatternFill("solid", fgColor="E5EDFB")
    for ci in range(1, len(headers) + 1):
        ws.cell(row_idx, ci).font = tot_font
        ws.cell(row_idx, ci).fill = tot_fill
    # 标签放第一个"非汇总"列,避免覆盖数值
    label_ci = 1
    for ci, h in enumerate(headers, 1):
        if h not in sum_cols and h not in mean_cols and h not in values:
            label_ci = ci
            break
    ws.cell(row_idx, label_ci, label)
    for ci, h in enumerate(headers, 1):
        if ci == label_ci:
            continue
        try:
            if h in values:
                val = float(values[h])                     # 调用方显式给的合计值最优先
            elif h in sum_cols and pd.api.types.is_numeric_dtype(df[h]):
                val = float(df[h].sum())
            elif h in mean_cols and pd.api.types.is_numeric_dtype(df[h]):
                w = _weighted_ratio_total(df, h)           # 比率列优先加权口径
                val = float(w) if w is not None else float(df[h].mean())
            else:
                continue
        except Exception:
            continue
        cell = ws.cell(row_idx, ci, val)
        cell.font = tot_font
        cell.fill = tot_fill
        if ci in col_fmt:
            cell.number_format = col_fmt[ci]


def export_cipher_as_is(
    cipher_path: Path,
    metadata_rows: list[dict],
    metadata_columns: list[str],
    stem: Optional[str] = None,
) -> Path:
    """
    用户在解密前就选「保留密文」(尚无计算结果)时的输出 —— 直接导出源密文:
      - 数值列保持源 cipher(base64),身份列(姓名/大区/月份)明文可见
      - 顶部一张「说明」sheet
    """
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dst = enforce_excel_path(make_excel_path(stem))
    dst.parent.mkdir(parents=True, exist_ok=True)

    try:
        cipher_df = pd.read_excel(cipher_path, index_col=0)
    except Exception:
        cipher_df = pd.read_excel(cipher_path)
    cipher_df = cipher_df.reset_index(drop=True)
    if metadata_rows and len(metadata_rows) == len(cipher_df):
        meta_df = pd.DataFrame(metadata_rows)
        keep = [c for c in (metadata_columns or list(meta_df.columns)) if c not in cipher_df.columns]
        if keep:
            cipher_df = pd.concat(
                [meta_df[keep].reset_index(drop=True), cipher_df], axis=1
            )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    notice = wb.create_sheet(title="说明")
    notice["A1"] = "保留密文 · 未授权解密"
    notice["A1"].font = Font(bold=True, size=13, color="9A3412")
    notice["A1"].fill = PatternFill("solid", fgColor="FFEDD5")
    for i, t in enumerate(
        ["", "用户选择保留密文,未解密 · 数值列保持同态密文形式。",
         "明文身份列(姓名 / 大区 / 月份 等)保留以便核对。",
         f"源密文文件:{cipher_path.name}"], 2):
        notice.cell(i, 1, t)
    notice.column_dimensions["A"].width = 80

    ws = wb.create_sheet(title="数据(密文保留)")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    headers = [str(c) for c in cipher_df.columns]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(cipher_df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, _defuse_formula(val) if isinstance(val, str) else val)
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(h) * 2.1, 12), 36)
    ws.freeze_panes = "A2"

    wb.save(dst)
    return dst


def export_skill_results_encrypted(
    results: list[dict],
    cipher_path: Path,
    note: str = "",
    stem: Optional[str] = None,
    staging: bool = False,
) -> Path:
    """
    用户选择「保留密文」时的输出 —— 与解密版同结构,但数值列再加密:
      - 多 sheet,每个 SkillCall 输出一张
      - 列布局同明文版(身份列在前,聚合 / 派生列在后)
      - 身份列(字符串)保持明文;数值列经 ct.encrypt_excel 再加密成 base64 cipher
      - 第一张「说明」sheet 列出已完成的 skill → sheet 映射
      - 无 chart(密文不可作图)
    """
    import openpyxl
    import pandas as pd
    import tempfile
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 初始化 HE runtime,获取 ct(用与密文文件同一套 sk/evk)
    from client.tools.runtime import Runtime
    Runtime.get().ensure_all_initialized()
    import crypto_toolkit as ct  # noqa: F401

    dst = enforce_excel_path(make_excel_path(stem, staging=staging))
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")

    # ---- 说明 sheet ----
    notice = wb.create_sheet(title="说明")
    notice["A1"] = "结果保留密文 · 未授权解密展示"
    notice["A1"].font = Font(bold=True, size=13, color="9A3412")
    notice["A1"].fill = PatternFill("solid", fgColor="FFEDD5")
    notice["A1"].alignment = Alignment(horizontal="left", vertical="center")
    lines = [
        "",
        "密态计算已在本机完成 · 全程未暴露明文。",
        "下方各 sheet 与解密版结构一致,但数值列保持同态密文(base64 ciphertext)形式。",
        "明文身份列(姓名 / 大区 / 月份 等)仍可见,便于核对结构。",
        f"源密文文件:{cipher_path.name}",
        "",
        "已完成的 skill → 结果 sheet:",
    ]
    for r in results:
        df = r.get("df")
        if df is None or df.empty:
            continue
        skill = r.get("skill", "?")
        nm = r.get("sheet_name", "?")
        lines.append(f"  · {skill} → 「{nm}」 ({len(df)} 行 × {len(df.columns)} 列)")
    if note:
        lines.extend(["", note])
    for i, t in enumerate(lines, 2):
        notice.cell(i, 1, t)
    notice.column_dimensions["A"].width = 80

    # ---- 每张结果 sheet ----
    for r in results:
        sheet_name = str(r.get("sheet_name") or "Sheet")[:31]
        df = r.get("df")
        if df is None or df.empty:
            continue

        # 行数护栏:再加密逐值很慢且会产出打不开的巨型文件 —— 超上限截断并在 note 标注
        enc_trunc = ""
        if len(df) > _ENC_ROW_LIMIT:
            enc_trunc = f"⚠ 结果 {len(df)} 行,密文版已截断前 {_ENC_ROW_LIMIT} 行。"
            df = df.head(_ENC_ROW_LIMIT)

        enc_df = _encrypt_numeric_columns(df)

        ws = wb.create_sheet(title=sheet_name)
        headers = [str(c) for c in enc_df.columns]
        # 口径说明行:与明文版一致地把 note(加权/等权口径、缺期预警等)写在表顶,
        # 避免"密文留存版看不到口径、明文版能看到"的两套结论
        header_row = 1
        sheet_note = (str(r.get("note") or "").strip() + " " + enc_trunc).strip()
        if sheet_note:
            c = ws.cell(1, 1, sheet_note)
            c.font = Font(italic=True, color="6B7280", size=10)
            if len(headers) > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            header_row = 2
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(header_row, ci, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for ri, row in enumerate(enc_df.itertuples(index=False), header_row + 1):
            for ci, val in enumerate(row, 1):
                cv = _defuse_formula(val) if isinstance(val, str) else val
                ws.cell(ri, ci, cv if cv is not None else "")
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = min(max(len(h) * 2.1, 12), 36)
        ws.freeze_panes = f"A{header_row + 1}"

    wb.save(dst)
    return dst


def _encrypt_numeric_columns(df):
    """
    对一张 skill 结果 df 做"再加密":
      身份列(非数值,如 string / category / object)保持明文原样;
      数值列(int / float)单独走 ct.encrypt_excel 文件管线 →
        读回来时数值单元格变成 base64 cipher 字符串。
      最终把两类列按原始列顺序拼回去,保持与解密版一致的布局。
    失败兜底:数值列填充 "[已加密]" 占位字符串,身份列原样保留。
    """
    import pandas as pd
    import tempfile

    try:
        from client.tools.runtime import Runtime
        Runtime.get().ensure_all_initialized()
        import crypto_toolkit as ct  # noqa: F401
    except Exception:
        ct = None

    original_cols = list(df.columns)
    df = df.copy().reset_index(drop=True)
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if not numeric_cols or ct is None:
        return df

    # 身份列(原样保留明文)
    id_cols = [c for c in original_cols if c not in numeric_cols]
    id_df = df[id_cols] if id_cols else None

    # 数值列单独写盘 → 加密 → 读回
    num_df = df[numeric_cols].copy().fillna(0)

    # 注:不能用 mkstemp()[1] —— 它会泄漏一个打开的 fd,Windows 上该句柄占着文件,
    # 随后 to_excel/encrypt_excel 写同一路径会报 WinError 32(文件被占用)。
    # 用 NamedTemporaryFile + with 立即关闭句柄(delete=False 保留文件)。
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _f:
        tmp_plain = Path(_f.name)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _f:
        tmp_enc = Path(_f.name)
    enc_num_df = None
    try:
        num_df.to_excel(str(tmp_plain), index=False)
        try:
            ct.encrypt_excel(str(tmp_plain), str(tmp_enc), input_index_col=None)
            # 加密 xlsx 里现在全是字符串(base64 cipher),用 dtype=str 读避免 pandas 自动 cast
            enc_num_df = pd.read_excel(str(tmp_enc), dtype=str)
        except Exception:
            enc_num_df = None
    finally:
        tmp_plain.unlink(missing_ok=True)
        tmp_enc.unlink(missing_ok=True)

    if enc_num_df is None:
        # 加密失败兜底
        enc_num_df = pd.DataFrame({c: ["[已加密]"] * len(df) for c in numeric_cols})

    # encrypt_excel 可能在末尾补 cipher metadata 行 —— 裁回原行数
    enc_num_df = enc_num_df.reset_index(drop=True).iloc[: len(df)].copy()

    # 数值列各 cell 已是 cipher(浮点的科学计数 / 或 NaN);为了 Excel 里可读,
    # 一律 toString 之后写入,且不能再被 Excel 当数字渲染。
    for c in enc_num_df.columns:
        enc_num_df[c] = enc_num_df[c].apply(
            lambda v: "" if pd.isna(v) else (v if isinstance(v, str) else f"{v:.16e}")
        )

    # 拼回去(按原始顺序)
    if id_df is not None:
        merged = pd.concat([id_df.reset_index(drop=True), enc_num_df], axis=1)
    else:
        merged = enc_num_df
    merged = merged[[c for c in original_cols if c in merged.columns]]
    return merged


def _time_gap_warning(df, chart_specs: list) -> str:
    """
    折线图 x 轴是时间列时检测断档:缺月的折线会把 1月直接连到 3月,看似平滑实为缺数。
    返回「⚠ 时间不连续:缺 2月、5月」式提示(轻量版,只提示不重构图);无断档返回 ""。
    """
    try:
        from client.tools.skills import _is_time_col, _time_sort_series
    except Exception:
        return ""
    for spec in chart_specs or []:
        if (spec.get("type") or "").lower() != "line":
            continue
        x = spec.get("x")
        if not x or x not in df.columns or not _is_time_col(x, df[x]):
            continue
        try:
            keys = {k for k in _time_sort_series(df[x]).tolist()
                    if isinstance(k, tuple) and all(v == v and v != float("inf") for v in k)}
            if not keys or len({len(k) for k in keys}) != 1:
                continue
            width = len(next(iter(keys)))
            # 归一成线性序号:纯月份=月号;(年,月)=年*12+月;其余粒度只查同前缀内的断档
            if width == 1:
                nums = sorted(k[0] for k in keys)
            elif width == 2:
                nums = sorted(k[0] * 12 + k[1] for k in keys)
            else:
                continue
            missing = [n for n in range(nums[0] + 1, nums[-1]) if n not in set(nums)]
            if not missing:
                continue
            if width == 1:
                names = [f"{m}月" for m in missing[:6]]
            else:
                names = [f"{(m - 1) // 12}年{(m - 1) % 12 + 1}月" for m in missing[:6]]
            more = f" 等 {len(missing)} 期" if len(missing) > 6 else ""
            return f"⚠ 时间不连续:缺 {'、'.join(names)}{more},折线在缺口处为跨期直连"
        except Exception:
            continue
    return ""


def _collect_chart_specs(r: dict) -> list:
    """从 result dict 收集图表规格 —— 支持单个 chart 或 charts 列表(多图)。"""
    specs: list = []
    one = r.get("chart")
    if isinstance(one, dict):
        specs.append(one)
    many = r.get("charts")
    if isinstance(many, (list, tuple)):
        specs.extend([c for c in many if isinstance(c, dict)])
    return specs


_MAX_SPLIT_CHARTS = 12    # 一个 spec 最多出多少张图(split_by 分组 + 分页合计),防爆炸
_MAX_ROWS_PER_CHART = 30  # 单张图最多画多少行(类别);超过则按连续区间分页,每页一张图


def _render_charts(ws, df, specs: list, headers: list, header_row: int, n_cols: int) -> None:
    """
    渲染所有图表,竖直堆叠在数据右侧。
    每个 spec 可带 split_by="<列名>" —— 按该列把数据拆成每组一张图
    (如 100 人按"销售大区"拆成 6 张、两个产品各出一张趋势图),
    每张图都有完整标题 + 横纵轴。df 已在上游按 split_by 预排序 → 同组行连续。
    """
    from openpyxl.utils import get_column_letter

    anchor_col = get_column_letter(n_cols + 2)
    data0 = header_row + 1
    anchor_row = header_row
    for spec in specs:
        for chart in _expand_chart_spec(ws, df, spec, headers, data0):
            if chart is None:
                continue
            ws.add_chart(chart, f"{anchor_col}{anchor_row}")
            anchor_row += 18   # 每张图竖直间隔约 18 行,互不遮挡


# 图表里要剔除的"求和总数"行 —— x 轴类别出现这些词的行不画进图
# (数据表里保留;只是不让一根"合计"柱把其余类别压扁)
_CHART_TOTAL_KEYS = ("合计", "总计", "汇总", "累计", "总和", "total")

# 自动配图时不当 y 的"序数列" / 不当 x 的"标识列"
_AUTO_Y_SKIP = ("排名", "名次", "序号")
_AUTO_X_SKIP = ("序号", "编号", "ID", "id", "备注", "说明")


def _auto_chart_spec(df) -> Optional[dict]:
    """
    一张 sheet 没有任何有效图表提示时的兜底自动配图:
      x = 基数最高的文本列(明细主体:销售代表/物料名称…,跳过序号/编号/备注)
      y = 第一个指标数值列(跳过排名/序号)
      行数 > 20 → 自动找低基数文本列(2~12 组,如销售大区)做 split_by 拆图
    挑不出有意义的列(纯数值表/相关性矩阵等)→ 返回 None,不强行画。
    """
    import pandas as pd
    try:
        if df is None or len(df) == 0:
            return None
        num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        txt_cols = [c for c in df.columns if c not in num_cols]
        y = next((c for c in num_cols
                  if not any(k in str(c) for k in _AUTO_Y_SKIP)), None)
        x_cands = [c for c in txt_cols
                   if not any(k in str(c) for k in _AUTO_X_SKIP)]
        if y is None or not x_cands:
            return None
        x = max(x_cands, key=lambda c: df[c].astype(str).nunique())
        spec = {"type": "bar", "x": str(x), "y": str(y), "title": f"{y} · 概览"}
        if len(df) > 20:
            cand = [(c, df[c].astype(str).nunique()) for c in x_cands if c != x]
            cand = [(c, n) for c, n in cand if 2 <= n <= 12]
            if cand:
                spec["split_by"] = str(min(cand, key=lambda t: t[1])[0])
        return spec
    except Exception:
        return None


def _is_total_label(v) -> bool:
    s = str(v).strip().lower()
    return bool(s) and any(k in s for k in _CHART_TOTAL_KEYS)


def _trim_trailing_totals(df, x_name: str) -> int:
    """返回剔除末尾「合计/总计…」行后的有效行数(合计行通常在表尾)。"""
    n = len(df)
    try:
        if x_name in df.columns:
            vals = df[x_name].astype(str).tolist()
            while n > 0 and _is_total_label(vals[n - 1]):
                n -= 1
    except Exception:
        pass
    return n


def _expand_chart_spec(ws, df, spec: dict, headers: list, data0: int) -> list:
    """一个 spec → 一张或多张 chart(有 split_by 时按组各一张)。"""
    typ = (spec.get("type") or "bar").lower()
    x = spec.get("x")
    y = spec.get("y")
    base_title = spec.get("title") or ""
    split_by = spec.get("split_by")

    if not (x and y) or x not in headers:
        return []
    y_cols = [y] if isinstance(y, str) else list(y)
    valid_y = [c for c in y_cols if c in headers]
    if not valid_y:
        return []
    x_idx = headers.index(x) + 1
    y_idxs = [headers.index(c) + 1 for c in valid_y]

    # 量纲检测:金额(百万级)和比率(0~1)画同一张图会把比率压成地平线。
    # 多 y 系列量纲差 >50 倍时,小量纲系列放**次轴**(折线),各自有自己的刻度。
    sec_y_idxs: list = []
    if typ in ("bar", "line") and len(valid_y) >= 2:
        try:
            import pandas as pd
            scales = {c: float(pd.to_numeric(df[c], errors="coerce").abs().max())
                      for c in valid_y}
            big = max(v for v in scales.values() if v == v)  # 忽略 NaN
            small = [c for c, v in scales.items() if v > 0 and big / v > 50]
            if small and len(small) < len(valid_y):
                sec_y_idxs = [headers.index(c) + 1 for c in small]
        except Exception:
            sec_y_idxs = []

    charts: list = []
    if split_by and split_by in df.columns:
        # df 已按 split_by 预排序 → 每组是连续行块
        col = df[split_by].astype(str).tolist()
        spans = []
        start = 0
        for j in range(1, len(col) + 1):
            if j == len(col) or col[j] != col[start]:
                spans.append((col[start], start, j - 1))
                start = j
        for gval, s, e in spans:
            if len(charts) >= _MAX_SPLIT_CHARTS:
                break
            if _is_total_label(gval):
                continue  # 「合计」组不单独出图
            title = f"{gval} · {base_title}" if base_title else str(gval)
            for fr, lr, page_title in _paginate_rows(data0 + s, data0 + e, title):
                if len(charts) >= _MAX_SPLIT_CHARTS:
                    break
                charts.append(_make_chart(ws, typ, x_idx, y_idxs, headers,
                                          fr, lr, page_title, sec_y_idxs))
    else:
        eff = _trim_trailing_totals(df, x)
        if eff <= 0:
            return []
        blocks = _entity_block_starts(df, x, eff)   # 分页对齐实体块,不拦腰切断
        for fr, lr, page_title in _paginate_rows(data0, data0 + eff - 1, base_title, blocks):
            charts.append(_make_chart(ws, typ, x_idx, y_idxs, headers,
                                      fr, lr, page_title, sec_y_idxs))
    return charts


def _entity_block_starts(df, x_name: str, n_rows: int) -> list:
    """
    找实体块边界(0-based 起始偏移列表,首元素为 0)。
    实体块 = 某文本列里连续同值的行段(逐行明细通常已按产品/大区聚块)。
    选块数最少(最外层维度)的文本列;全同值/全唯一的列不算块维度,找不到返回 []。
    """
    import pandas as pd
    best = None
    for c in df.columns:
        if str(c) == str(x_name):
            continue
        s = df[c].iloc[:n_rows]
        if pd.api.types.is_numeric_dtype(s):
            continue
        vals = s.astype(str).tolist()
        starts = [0] + [i for i in range(1, n_rows) if vals[i] != vals[i - 1]]
        n_runs = len(starts)
        if n_runs <= 1 or n_runs >= n_rows:
            continue
        if best is None or n_runs < best[0]:
            best = (n_runs, starts)
    return best[1] if best else []


def _paginate_rows(first_row: int, last_row: int, base_title: str,
                   block_starts: Optional[list] = None) -> list:
    """
    把 [first_row, last_row] 行区间按 _MAX_ROWS_PER_CHART 分页,每页一张图,
    标题带「(第 i/n 页 · 第 a-b 行)」后缀。页数封顶 _MAX_SPLIT_CHARTS
    (超出时自动加大每页行数,保证所有数据仍被画全)。

    传入 block_starts(实体块 0-based 起始偏移)时,页边界对齐块边界——
    整块整块装页,不把某个产品的时间序列拦腰切到两张图;单块超页容量才按行切。
    """
    import math

    n = last_row - first_row + 1
    if n <= _MAX_ROWS_PER_CHART:
        return [(first_row, last_row, base_title)]
    size = max(_MAX_ROWS_PER_CHART, math.ceil(n / _MAX_SPLIT_CHARTS))

    def _spans_by_rows(sz: int) -> list:
        pages = math.ceil(n / sz)
        sz = math.ceil(n / pages)   # 均匀化
        return [(i * sz, min(n - 1, (i + 1) * sz - 1)) for i in range(pages)]

    def _spans_by_blocks(sz: int) -> list:
        bounds = list(block_starts) + [n]
        spans, cur_s, cur_len = [], 0, 0
        for bi in range(len(bounds) - 1):
            b0, b1 = bounds[bi], bounds[bi + 1]
            blk = b1 - b0
            if blk > sz:                       # 单块超页容量 → 先收尾,再块内按行切
                if cur_len:
                    spans.append((cur_s, b0 - 1))
                spans.extend((b0 + s, b0 + e) for s, e in
                             [(i * sz, min(blk - 1, (i + 1) * sz - 1))
                              for i in range(math.ceil(blk / sz))])
                cur_s, cur_len = b1, 0
            elif cur_len and cur_len + blk > sz:  # 装不下 → 换页
                spans.append((cur_s, b0 - 1))
                cur_s, cur_len = b0, blk
            else:
                if not cur_len:
                    cur_s = b0
                cur_len += blk
        if cur_len:
            spans.append((cur_s, n - 1))
        return spans

    if block_starts:
        spans = _spans_by_blocks(size)
        tries = 0
        while len(spans) > _MAX_SPLIT_CHARTS and tries < 4:   # 页数超限 → 加大页容量重排
            size = math.ceil(size * 1.5)
            spans = _spans_by_blocks(size)
            tries += 1
        if len(spans) > _MAX_SPLIT_CHARTS:
            spans = _spans_by_rows(size)
    else:
        spans = _spans_by_rows(size)

    out = []
    for i, (s, e) in enumerate(spans):
        suffix = f"(第 {i + 1}/{len(spans)} 页 · 第 {s + 1}-{e + 1} 行)"
        out.append((first_row + s, first_row + e,
                    f"{base_title} {suffix}" if base_title else suffix))
    return out


def _make_chart(ws, typ: str, x_idx: int, y_idxs: list, headers: list,
                first_row: int, last_row: int, title: str,
                sec_y_idxs: Optional[list] = None):
    """
    构建一张产品级图表(支持任意连续行子区间):
    显式系列名(不依赖表头连续)/ 轴标题 / 图例底部 / 折线圆点 / Y 轴数字格式 / 网格。
    类型:bar(默认) / line / pie(占比) / stacked_bar(构成堆积)。
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series

    if last_row < first_row:
        return None
    if typ == "pie":
        chart = PieChart()
    elif typ == "line":
        chart = LineChart()
    else:
        chart = BarChart()
        if typ in ("stacked", "stacked_bar"):
            chart.type = "col"
            chart.grouping = "stacked"
            chart.overlap = 100
    chart.title = title or None
    chart.style = 12
    chart.height = 8.5
    chart.width = 20

    cats = Reference(ws, min_col=x_idx, min_row=first_row, max_row=last_row)
    sec = set(sec_y_idxs or []) if typ != "pie" else set()
    use_y = y_idxs[:1] if typ == "pie" else [yi for yi in y_idxs if yi not in sec]
    if not use_y and y_idxs:   # 全被划去次轴时至少留一个在主轴
        use_y, sec = y_idxs[:1], sec - {y_idxs[0]}
    for yi in use_y:
        data = Reference(ws, min_col=yi, min_row=first_row, max_row=last_row)
        chart.series.append(Series(data, title=headers[yi - 1]))  # 显式系列名
    chart.set_categories(cats)
    if sec:
        # 次轴组合图:小量纲系列(比率等)画成折线挂在右侧第二 Y 轴
        try:
            sub = LineChart()
            for yi in sorted(sec):
                data = Reference(ws, min_col=yi, min_row=first_row, max_row=last_row)
                sub.series.append(Series(data, title=headers[yi - 1]))
            sub.y_axis.axId = 200
            sub.y_axis.title = headers[sorted(sec)[0] - 1]
            sub.y_axis.crosses = "max"
            sub.y_axis.numFmt = _infer_number_format(headers[sorted(sec)[0] - 1]) or "0.00%"
            from openpyxl.chart.marker import Marker
            for s_ in sub.series:
                s_.marker = Marker(symbol="circle", size=5)
            chart += sub
        except Exception:
            pass
    if typ == "pie":
        # 饼图无 x/y 轴:单独设图例底部 + 百分比标签,不进下面的轴配置块(否则 AttributeError)
        try:
            from openpyxl.chart.label import DataLabelList
            chart.dataLabels = DataLabelList(showPercent=True)   # 占比直接标在饼上
            chart.legend.position = "b"
        except Exception:
            pass
        return chart

    try:
        chart.x_axis.title = headers[x_idx - 1]
        chart.y_axis.title = headers[y_idxs[0] - 1] if len(y_idxs) == 1 else None
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        from openpyxl.chart.axis import ChartLines
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.numFmt = _infer_number_format(headers[y_idxs[0] - 1]) or "General"
        chart.legend.position = "b"
        # 长类目名(如产品全称)斜排 -45°,避免 x 轴标签挤成一团
        try:
            cats_max = max((len(str(ws.cell(r, x_idx).value or ""))
                            for r in range(first_row, last_row + 1)), default=0)
            if cats_max >= 6:
                chart.x_axis.txPr = _rotated_text_props(-45)
        except Exception:
            pass
    except Exception:
        pass

    if typ == "line":
        try:
            from openpyxl.chart.marker import Marker
            for s in chart.series:
                s.marker = Marker(symbol="circle", size=5)
                s.smooth = False
        except Exception:
            pass

    return chart
