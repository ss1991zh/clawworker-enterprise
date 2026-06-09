"""
密态定时任务的结果暂存 / 批量解密。

定时密态任务到点正常计算,但结果**不解密**,而是加密暂存到沙盒;
用户回来按任务批量解密 → 所有结果明文 Excel 落到一个文件夹。

每张结果 df 拆两部分持久化(沙盒 ~/.agent-system/scheduler/enc_results/):
  - 数值列  → ct.encrypt_excel 暂存为可再解密的密文 xlsx
  - 身份列  → 明文 csv sidecar(姓名 / 大区 等,本系统一贯按明文 metadata 处理)
  - manifest 记录 sheet 名 + 两个文件路径 + 列顺序

批量解密:ps.read_excel + ct.decrypt_df 还原数值列 → 拼回身份列 → 多 sheet 明文 Excel。
"""

from __future__ import annotations

import json
import secrets
from pathlib import Path
from typing import Any, Optional


ENC_RESULTS_DIR = Path.home() / ".agent-system" / "scheduler" / "enc_results"


def persist_results_encrypted(results: list[dict], run_id: str) -> list[dict]:
    """
    把一次运行的 results([{sheet_name, df, chart}])加密暂存到沙盒。
    返回 manifest 列表:[{sheet_name, num_enc, id_csv, col_order, n_rows}]
    """
    import pandas as pd
    import tempfile

    from client.tools.runtime import Runtime
    Runtime.get().ensure_all_initialized()
    import crypto_toolkit as ct  # noqa: F401

    run_dir = ENC_RESULTS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict] = []
    for idx, r in enumerate(results):
        df = r.get("df")
        if df is None or getattr(df, "empty", True):
            continue
        df = df.copy().reset_index(drop=True)
        col_order = [str(c) for c in df.columns]
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        id_cols = [c for c in df.columns if c not in numeric_cols]

        num_enc_path = ""
        if numeric_cols:
            num_df = df[numeric_cols].copy().fillna(0)
            tmp_plain = Path(tempfile.mkstemp(suffix=".xlsx")[1])
            enc_path = run_dir / f"sheet{idx}_num.xlsx"
            try:
                num_df.to_excel(str(tmp_plain), index=False)
                ct.encrypt_excel(str(tmp_plain), str(enc_path), input_index_col=None)
                num_enc_path = str(enc_path)
            finally:
                tmp_plain.unlink(missing_ok=True)

        id_csv_path = ""
        if id_cols:
            id_df = df[id_cols].copy()
            p = run_dir / f"sheet{idx}_id.csv"
            id_df.to_csv(str(p), index=False)
            id_csv_path = str(p)

        manifest.append({
            "sheet_name": str(r.get("sheet_name") or f"结果{idx+1}")[:31],
            "num_enc": num_enc_path,
            "id_csv": id_csv_path,
            "numeric_cols": [str(c) for c in numeric_cols],
            "col_order": col_order,
            "n_rows": int(len(df)),
        })

    # manifest 也落一份(方便排查)
    (run_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def _decrypt_one_sheet(entry: dict):
    """按 manifest 单条还原一张明文 df。"""
    import pandas as pd

    from client.tools.runtime import Runtime
    Runtime.get().ensure_all_initialized()
    import crypto_toolkit as ct  # noqa: F401
    import pandaseal as ps  # noqa: F401

    num_plain = None
    if entry.get("num_enc") and Path(entry["num_enc"]).exists():
        try:
            cdf = ps.read_excel(entry["num_enc"], index_col=0)
        except Exception:
            cdf = ps.read_excel(entry["num_enc"])
        num_plain = ct.decrypt_df(cdf).reset_index(drop=True)
        # 列名对齐(encrypt 往返后列名应保持;兜底用 numeric_cols)
        if list(num_plain.columns) != entry.get("numeric_cols", []):
            try:
                num_plain.columns = entry["numeric_cols"][: len(num_plain.columns)]
            except Exception:
                pass

    id_plain = None
    if entry.get("id_csv") and Path(entry["id_csv"]).exists():
        id_plain = pd.read_csv(entry["id_csv"]).reset_index(drop=True)

    if id_plain is not None and num_plain is not None:
        merged = pd.concat([id_plain, num_plain], axis=1)
    elif num_plain is not None:
        merged = num_plain
    elif id_plain is not None:
        merged = id_plain
    else:
        merged = pd.DataFrame()

    # 还原原始列顺序
    order = [c for c in entry.get("col_order", []) if c in merged.columns]
    if order:
        merged = merged[order]
    return merged


def decrypt_runs_to_folder(runs: list[dict], folder_name: str) -> Path:
    """
    批量解密多次运行 → 明文 Excel 落到 ~/Downloads/<folder_name>/。

    runs: [{run_id, run_at, manifest: [...], question}]
      每次运行 → 一个 Excel 文件(多 sheet)。
    返回文件夹路径。
    """
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from client.webui.writer import _infer_number_format

    downloads = Path.home() / "Downloads"
    safe = "".join(ch for ch in folder_name if ch not in '\\/:*?"<>|').strip() or "定时任务结果"
    out_dir = downloads / safe
    # 防重名
    if out_dir.exists():
        out_dir = downloads / f"{safe}_{datetime.now().strftime('%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")

    for run in runs:
        manifest = run.get("manifest", []) or []
        ts = (run.get("run_at") or "").replace(":", "").replace("-", "").replace("T", "_")[:15]
        fname = f"{safe}_{ts or secrets.token_hex(3)}.xlsx"
        dst = out_dir / fname

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        any_sheet = False
        for entry in manifest:
            df = _decrypt_one_sheet(entry)
            if df is None or df.empty:
                continue
            any_sheet = True
            ws = wb.create_sheet(title=str(entry.get("sheet_name") or "结果")[:31])
            headers = [str(c) for c in df.columns]
            col_fmt = {}
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(1, ci, h)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                nf = _infer_number_format(h)
                if nf:
                    col_fmt[ci] = nf
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(ri, ci, val)
                    if ci in col_fmt:
                        cell.number_format = col_fmt[ci]
            for ci, h in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(ci)].width = min(max(len(h) * 2.1, 12), 30)
            ws.freeze_panes = "A2"
        if any_sheet:
            wb.save(dst)

    return out_dir


def cleanup_runs(run_ids: list[str]) -> None:
    """解密完成后删掉沙盒里的密文暂存。"""
    import shutil
    for rid in run_ids:
        d = ENC_RESULTS_DIR / rid
        try:
            if d.is_dir():
                shutil.rmtree(d, ignore_errors=True)
        except Exception:
            pass
