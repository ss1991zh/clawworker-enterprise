"""
B5 Excel 输出 —— 用户的最终交付物(architecture.md §B5)。

约束:
- 路径必须在 ~/Downloads/(通过 permissions.enforce_excel_path 强制)
- 文件名 analysis_<YYYYMMDD_HHMMSS>.xlsx,每次新文件不覆盖
- 图表用 openpyxl 在 Excel 内原生生成
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.series import Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from client.permissions import DOWNLOADS_DIR, enforce_excel_path
from shared.contract import ChartSpec, ChartType


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


@dataclass
class KpiCard:
    """顶部 KPI 卡片(label 小字 + value 大字 + subtitle 副标题)。"""

    label: str
    value: str
    subtitle: Optional[str] = None
    bg_color: str = "EAF1F8"
    value_color: str = "1F4E79"
    value_size: int = 16


@dataclass
class SheetData:
    """一个 sheet 的实际内容:表头 + 行 + 可选图表 + 格式 hint。

    格式与样式 hints(可选,不传则不应用):
    - number_formats:{列名: openpyxl number_format}(如 "0.00%")
    - tier_colors:{列名: [(upper_bound, hex_color), ...]} —— 按值落入区间涂色
    - extra_category_col:柱状图主 X 轴外再叠加一层分组列名(分层 X 轴)
    - kpi_cards:顶部 KPI 卡片;若有则 headers 自动下移
    - series_colors_by_row:与 rows 等长的 hex 色列表;chart 第 1 个 series 的每个 bar 单独染色
    """

    name: str
    headers: list[str]
    rows: list[list[Any]]
    chart: Optional[ChartSpec] = None
    number_formats: dict[str, str] = field(default_factory=dict)
    tier_colors: dict[str, list[tuple[float, str]]] = field(default_factory=dict)
    extra_category_col: Optional[str] = None
    kpi_cards: list[KpiCard] = field(default_factory=list)
    series_colors_by_row: Optional[list[str]] = None


# KPI 区域占 4 行(1 label / 2-3 value / 4 subtitle),+1 行空白分隔
KPI_BAND_ROWS = 4
KPI_BAND_SEPARATOR = 1
KPI_CARD_COLS = 3  # 每张卡片占 3 列


@dataclass
class WriteResult:
    """Excel 写入结果。"""

    path: Path
    sheet_names: list[str] = field(default_factory=list)
    charts_generated: int = 0


# ---------------------------------------------------------------------------
# 文件名生成
# ---------------------------------------------------------------------------


def make_excel_path(prefix: str = "analysis", now: Optional[datetime] = None) -> Path:
    """
    生成带时间戳的 Excel 文件路径,确保不与既存文件冲突。
    返回 ~/Downloads/<prefix>_<YYYYMMDD_HHMMSS>[_<n>].xlsx 形式。
    """
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    ts = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    candidate = DOWNLOADS_DIR / f"{prefix}_{ts}.xlsx"
    n = 1
    # 若同秒生成多个文件,加 _1 / _2 ...
    while candidate.exists():
        candidate = DOWNLOADS_DIR / f"{prefix}_{ts}_{n}.xlsx"
        n += 1
    return candidate


# ---------------------------------------------------------------------------
# 图表生成
# ---------------------------------------------------------------------------


def _create_chart(
    chart_spec: ChartSpec,
    sheet,
    header_row: int = 1,
    extra_category_col: Optional[str] = None,
    series_colors_by_row: Optional[list[str]] = None,
    data_end_row: Optional[int] = None,
) -> Optional[Any]:
    """根据 ChartSpec 创建 openpyxl chart 对象并绑定数据。

    extra_category_col:若提供且与 chart_spec.x 在表中是连续两列(顺序无关),
    会用多列 categories 产生 Excel 分层 X 轴(如 大区 / 销售代表)。
    series_colors_by_row:与数据行 1-1 对齐的 hex 色列表 —— 给第 1 个 series 的每根 bar 单独染色
    data_end_row:显式数据末尾行(默认 sheet.max_row)
    """
    max_row = data_end_row or sheet.max_row
    max_col = sheet.max_column
    if max_row <= header_row:
        return None

    headers = [sheet.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]

    def col_idx(name: str) -> Optional[int]:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    x_col = col_idx(chart_spec.x)
    if x_col is None:
        return None

    y_cols: list[int] = []
    y_names = chart_spec.y if isinstance(chart_spec.y, list) else [chart_spec.y]
    for y_name in y_names:
        idx = col_idx(y_name)
        if idx is not None:
            y_cols.append(idx)
    if not y_cols:
        return None

    chart_type: ChartType = chart_spec.type
    chart_cls = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "heatmap": BarChart,
    }.get(chart_type, LineChart)

    chart = chart_cls()
    if chart_spec.title:
        chart.title = chart_spec.title

    for y_col in y_cols:
        data_ref = Reference(
            sheet,
            min_col=y_col,
            max_col=y_col,
            min_row=header_row,
            max_row=max_row,
        )
        chart.add_data(data_ref, titles_from_data=True)

    # 分层 X 轴:若 extra_category_col 是与 x 相邻的列,categories 用两列连片
    extra_col_idx = col_idx(extra_category_col) if extra_category_col else None
    if extra_col_idx is not None and abs(extra_col_idx - x_col) == 1:
        min_c, max_c = sorted([extra_col_idx, x_col])
        cat_ref = Reference(
            sheet,
            min_col=min_c,
            max_col=max_c,
            min_row=header_row + 1,
            max_row=max_row,
        )
    else:
        cat_ref = Reference(
            sheet,
            min_col=x_col,
            max_col=x_col,
            min_row=header_row + 1,
            max_row=max_row,
        )
    chart.set_categories(cat_ref)

    # 逐 bar 染色(仅作用于第 1 个 series)
    if series_colors_by_row and chart.series:
        s = chart.series[0]
        for i, color_hex in enumerate(series_colors_by_row):
            if not color_hex:
                continue
            pt = DataPoint(idx=i)
            pt.graphicalProperties = GraphicalProperties(solidFill=color_hex)
            s.dPt.append(pt)

    return chart


# ---------------------------------------------------------------------------
# 表头/格式/涂色辅助
# ---------------------------------------------------------------------------


HEADER_FILL = PatternFill("solid", fgColor="DDEEFF")
HEADER_FONT = Font(bold=True)


def _apply_number_formats(ws, sd: "SheetData", *, header_row: int = 1) -> None:
    data_start = header_row + 1
    for col_idx, header in enumerate(sd.headers, start=1):
        fmt = sd.number_formats.get(header)
        if not fmt:
            continue
        for r in range(data_start, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = fmt


def _apply_tier_colors(ws, sd: "SheetData", *, header_row: int = 1) -> None:
    data_start = header_row + 1
    for col_name, tiers in sd.tier_colors.items():
        if col_name not in sd.headers:
            continue
        col_idx = sd.headers.index(col_name) + 1
        for r in range(data_start, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if not isinstance(val, (int, float)):
                continue
            for upper, color in tiers:
                if val < upper:
                    ws.cell(row=r, column=col_idx).fill = PatternFill("solid", fgColor=color)
                    break


def _apply_header_style(ws, sd: "SheetData", *, header_row: int = 1) -> None:
    for c in range(1, len(sd.headers) + 1):
        ws.cell(header_row, c).font = HEADER_FONT
        ws.cell(header_row, c).fill = HEADER_FILL


# ---------------------------------------------------------------------------
# KPI 卡片渲染
# ---------------------------------------------------------------------------


def _write_kpi_band(ws, cards: list[KpiCard]) -> int:
    """
    在 sheet 顶部写一排 KPI 卡片。返回 header 应该写入的行号(KPI 占 4 行 + 1 行空白 = 6)。
    若 cards 为空,不写任何东西,返回 1。
    """
    if not cards:
        return 1

    for i, card in enumerate(cards):
        c0 = i * KPI_CARD_COLS + 1
        c1 = c0 + KPI_CARD_COLS - 1
        bg = PatternFill("solid", fgColor=card.bg_color)

        # label 行
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        cell = ws.cell(1, c0)
        cell.value = card.label
        cell.font = Font(size=10, color="666666")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = bg

        # value 行(2-3 合并,字大)
        ws.merge_cells(start_row=2, start_column=c0, end_row=3, end_column=c1)
        cell = ws.cell(2, c0)
        cell.value = card.value
        cell.font = Font(bold=True, size=card.value_size, color=card.value_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = bg

        # subtitle 行
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c1)
        cell = ws.cell(4, c0)
        cell.value = card.subtitle or ""
        cell.font = Font(size=9, color="666666", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = bg

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18
    return KPI_BAND_ROWS + KPI_BAND_SEPARATOR + 1  # = 6


# ---------------------------------------------------------------------------
# 主写入函数
# ---------------------------------------------------------------------------


def write_excel(
    sheets: list[SheetData],
    *,
    path: Optional[Path] = None,
) -> WriteResult:
    """
    把多 sheet 数据写入 Excel 文件。

    Args:
        sheets: 各 sheet 的内容
        path: 目标路径(必须在 ~/Downloads/),为 None 时自动生成

    Returns:
        WriteResult,含最终路径与生成的图表数
    """
    if not sheets:
        raise ValueError("至少需要一个 sheet")

    target = path or make_excel_path()
    target = enforce_excel_path(target)  # B6 第 2 条

    if target.exists():
        # 严格遵守"不覆盖旧文件"
        raise FileExistsError(f"目标文件已存在,不覆盖: {target}")

    wb = Workbook()
    # 默认 sheet 删掉,我们按 SheetData 重建
    default_ws = wb.active
    wb.remove(default_ws)

    charts_generated = 0
    sheet_names: list[str] = []

    for sd in sheets:
        ws = wb.create_sheet(title=sd.name[:31])
        sheet_names.append(ws.title)

        # KPI 卡片(若有)→ 返回 headers 的起始行号(默认 1,有卡片则 6)
        header_row = _write_kpi_band(ws, sd.kpi_cards)

        # 表头
        for col_idx, header in enumerate(sd.headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=header)

        # 数据行
        for row_offset, row in enumerate(sd.rows):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=header_row + 1 + row_offset, column=col_idx, value=value)
        data_end_row = header_row + len(sd.rows)

        # 自适应列宽
        for col_idx, header in enumerate(sd.headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) + 2)

        # 表头样式 / 列格式 / 档位涂色
        _apply_header_style(ws, sd, header_row=header_row)
        _apply_number_formats(ws, sd, header_row=header_row)
        _apply_tier_colors(ws, sd, header_row=header_row)

        # 图表
        if sd.chart:
            chart = _create_chart(
                sd.chart,
                ws,
                header_row=header_row,
                extra_category_col=sd.extra_category_col,
                series_colors_by_row=sd.series_colors_by_row,
                data_end_row=data_end_row,
            )
            if chart:
                anchor_col_idx = len(sd.headers) + 2
                anchor_col = get_column_letter(anchor_col_idx)
                ws.add_chart(chart, f"{anchor_col}{header_row}")
                charts_generated += 1

    wb.save(target)
    return WriteResult(path=target, sheet_names=sheet_names, charts_generated=charts_generated)
