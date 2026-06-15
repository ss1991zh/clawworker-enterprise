"""
定时任务批量解密回归测试 —— 锁住三类 bug:
  · 同秒完成的多次运行 → 文件名碰撞互相覆盖,只剩"最后一个 excel"
  · 某次运行密文缺失 → 静默跳过却仍被标记已解密 → 数据无声丢失
  · 单次运行失败拖垮整批
manifest 用纯身份列(id_csv),不依赖真实 HE。
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from client.webui.sched_results import decrypt_runs_to_folder


def _make_run(tmp_path: Path, run_id: str, run_at: str, rows: list[dict]) -> dict:
    """构造一个 id-only 的 run(无数值密文,manifest 仅身份列 csv)。"""
    df = pd.DataFrame(rows)
    csv = tmp_path / f"{run_id}_id.csv"
    df.to_csv(csv, index=False)
    return {
        "run_id": run_id, "run_at": run_at, "question": "测试",
        "manifest": [{
            "sheet_name": "结果",
            "num_enc": "", "id_csv": str(csv),
            "numeric_cols": [], "col_order": list(df.columns),
            "n_rows": len(df),
        }],
    }


def _cleanup(out_dir: Path):
    shutil.rmtree(out_dir, ignore_errors=True)


def test_each_run_gets_own_file(tmp_path: Path):
    runs = [
        _make_run(tmp_path, "run1", "2026-06-09T22:43:20", [{"大区": "华东"}]),
        _make_run(tmp_path, "run2", "2026-06-09T22:45:02", [{"大区": "华南"}]),
        _make_run(tmp_path, "run3", "2026-06-11T20:07:06", [{"大区": "华北"}]),
    ]
    out_dir, outcomes = decrypt_runs_to_folder(runs, "解密回归测试A")
    try:
        files = sorted(p.name for p in out_dir.glob("*.xlsx"))
        assert len(files) == 3, f"3 次运行应产出 3 个文件,实际 {files}"
        assert all(o["ok"] for o in outcomes)
        # 每个文件装自己的数据
        data = {}
        for p in out_dir.glob("*.xlsx"):
            ws = load_workbook(p)["结果"]
            data[p.name] = ws.cell(2, 1).value
        assert set(data.values()) == {"华东", "华南", "华北"}
    finally:
        _cleanup(out_dir)


def test_same_second_runs_do_not_overwrite(tmp_path: Path):
    """同一秒完成的两次运行 → 文件名去重,绝不互相覆盖(原 bug:只剩最后一个)。"""
    same_ts = "2026-06-11T20:08:32"
    runs = [
        _make_run(tmp_path, "runA", same_ts, [{"大区": "甲"}]),
        _make_run(tmp_path, "runB", same_ts, [{"大区": "乙"}]),
    ]
    out_dir, outcomes = decrypt_runs_to_folder(runs, "解密回归测试B")
    try:
        files = list(out_dir.glob("*.xlsx"))
        assert len(files) == 2, f"同秒 2 次运行应产出 2 个文件,实际只剩 {[p.name for p in files]}"
        vals = {load_workbook(p)["结果"].cell(2, 1).value for p in files}
        assert vals == {"甲", "乙"}, "两个文件应各装各的数据"
    finally:
        _cleanup(out_dir)


def test_missing_cipher_reported_not_silently_dropped(tmp_path: Path):
    """密文暂存缺失的 run → ok=False 报告出来,不产文件;好的 run 不受影响。"""
    good = _make_run(tmp_path, "good", "2026-06-11T20:07:06", [{"大区": "华东"}])
    bad = {
        "run_id": "bad", "run_at": "2026-06-09T22:45:02", "question": "测试",
        "manifest": [{
            "sheet_name": "结果",
            "num_enc": str(tmp_path / "不存在_num.xlsx"),   # 沙盒文件已丢失
            "id_csv": str(tmp_path / "不存在_id.csv"),
            "numeric_cols": ["金额"], "col_order": ["金额"], "n_rows": 7,
        }],
    }
    out_dir, outcomes = decrypt_runs_to_folder([bad, good], "解密回归测试C")
    try:
        by_id = {o["run_id"]: o for o in outcomes}
        assert by_id["good"]["ok"] is True and by_id["good"]["file"]
        assert by_id["bad"]["ok"] is False and by_id["bad"]["error"], \
            "密文缺失必须显式报告,不能静默跳过"
        assert len(list(out_dir.glob("*.xlsx"))) == 1
    finally:
        _cleanup(out_dir)


def test_presentation_hints_roundtrip(tmp_path: Path):
    """暂存的 chart/档位等呈现键 → 解密后用产品级渲染器还原(图表/样式与单独提问一致)。"""
    df = pd.DataFrame({"大区": ["华东", "华南", "华北"], "排名": [1, 2, 3]})
    csv = tmp_path / "p_id.csv"
    df.to_csv(csv, index=False)
    run = {
        "run_id": "p1", "run_at": "2026-06-11T20:49:31", "question": "回款率",
        "manifest": [{
            "sheet_name": "回款率排名",
            "num_enc": "", "id_csv": str(csv),
            "numeric_cols": [], "col_order": ["大区", "排名"], "n_rows": 3,
            "chart": {"type": "bar", "x": "大区", "y": "排名", "title": "回款率排名"},
            "total_row": False,
        }],
    }
    out_dir, outcomes = decrypt_runs_to_folder([run], "解密回归测试E")
    try:
        assert outcomes[0]["ok"]
        p = next(out_dir.glob("*.xlsx"))
        ws = load_workbook(p)["回款率排名"]
        assert len(ws._charts) == 1, "暂存的 chart 应在解密 Excel 里还原"
        assert ws.auto_filter.ref, "应有产品级渲染的自动筛选"
    finally:
        _cleanup(out_dir)


def test_all_failed_returns_no_folder_litter(tmp_path: Path):
    bad = {
        "run_id": "bad", "run_at": "2026-06-09T22:45:02", "question": "",
        "manifest": [{"sheet_name": "结果", "num_enc": str(tmp_path / "无.xlsx"),
                      "id_csv": "", "numeric_cols": ["x"], "col_order": ["x"], "n_rows": 1}],
    }
    out_dir, outcomes = decrypt_runs_to_folder([bad], "解密回归测试D")
    assert not any(o["ok"] for o in outcomes)
    assert not out_dir.exists(), "全部失败时不应留下空文件夹"
