"""
图表渲染回归测试 —— 锁住"可看性"能力:
  · split_by 按组拆成多张图(治"100人挤一图")
  · charts 列表 → 一张表多张图
  · 默认单图
纯 pandas + openpyxl,不依赖真实 HE。
"""
from __future__ import annotations

import pandas as pd

from client.webui import writer as W


def _load(results):
    from openpyxl import load_workbook
    out = W.write_skill_results(results, stem="chart_test")
    wb = load_workbook(out)
    out.unlink(missing_ok=True)
    return wb


def test_single_chart_default():
    df = pd.DataFrame({"区域": ["A", "B", "C"], "销量": [3, 1, 2]})
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "bar", "x": "区域", "y": "销量", "title": "T"}}])
    assert len(wb["S"]._charts) == 1


def test_split_by_one_chart_per_group():
    # 6 个大区 × 多人 → split_by 应产出 6 张图
    regions = ["华东", "华南", "华北", "西南", "华中", "海外"]
    rows = [{"代表": f"人{i:02d}", "大区": regions[i % 6], "贡献率": 0.1 + (i % 13) / 100}
            for i in range(60)]
    df = pd.DataFrame(rows)
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "bar", "x": "代表", "y": "贡献率",
                           "title": "个人贡献率", "split_by": "大区"}}])
    assert len(wb["S"]._charts) == 6


def test_split_by_sorts_table_by_group():
    df = pd.DataFrame({"代表": ["x", "y", "z"], "大区": ["华南", "华东", "华南"], "v": [1, 2, 3]})
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "bar", "x": "代表", "y": "v", "split_by": "大区"}}])
    ws = wb["S"]
    # 表格应按大区分组(同组连续):第一列数据区的大区列
    col_region = [ws.cell(r, 2).value for r in range(2, 5)]
    assert col_region == sorted(col_region)  # 已分组排序


def test_charts_list_multiple():
    df = pd.DataFrame({"区域": ["A", "B"], "销量": [3, 1], "利润": [2, 4]})
    wb = _load([{"sheet_name": "S", "df": df, "charts": [
        {"type": "bar", "x": "区域", "y": "销量", "title": "销量"},
        {"type": "bar", "x": "区域", "y": "利润", "title": "利润"},
    ]}])
    assert len(wb["S"]._charts) == 2


def test_multi_entity_forecast_two_line_charts():
    # 两个产品各历史+预测 → split_by 产品 → 2 张折线图,各 2 系列
    def rows(name, base):
        h = [{"产品": name, "月": f"m{m}", "历史值": float(base + m), "预测值": None} for m in range(1, 4)]
        f = [{"产品": name, "月": f"m{m}", "历史值": None, "预测值": float(base + m)} for m in range(4, 7)]
        return h + f
    df = pd.DataFrame(rows("DR-400", 100) + rows("SM-200", 50))
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "line", "x": "月", "y": ["历史值", "预测值"],
                           "title": "趋势", "split_by": "产品"}}])
    charts = wb["S"]._charts
    assert len(charts) == 2
    assert all(len(c.series) == 2 for c in charts)


def test_chart_excludes_trailing_total_row():
    """数据表末尾的「合计」行保留在表里,但不画进图表。"""
    df = pd.DataFrame({"区域": ["A", "B", "C", "合计"], "销量": [3, 1, 2, 6]})
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "bar", "x": "区域", "y": "销量", "title": "T"}}])
    ws = wb["S"]
    c = ws._charts[0]
    # 类别引用应止于第 4 行(表头1 + 数据A/B/C),不含第 5 行的「合计」
    cats_ref = c.series[0].cat.numRef or c.series[0].cat.strRef
    assert cats_ref.f.endswith("$4"), f"图表类别应剔除合计行: {cats_ref.f}"
    # 表格里合计行仍在
    assert ws.cell(5, 1).value == "合计"


def test_split_by_skips_total_group():
    df = pd.DataFrame({"代表": ["x", "y", "z"], "大区": ["华东", "华南", "总计"], "v": [1, 2, 3]})
    wb = _load([{"sheet_name": "S", "df": df,
                 "chart": {"type": "bar", "x": "代表", "y": "v", "split_by": "大区"}}])
    assert len(wb["S"]._charts) == 2, "「总计」组不应单独出图"


def test_auto_chart_when_no_hint():
    """LLM 漏配 chart → 渲染端兜底自动配图;行多自动按低基数维度拆。"""
    regions = ["华东", "华南", "华北", "西南", "华中", "东北"]
    df = pd.DataFrame([{"序号": str(i), "销售代表": f"代表{i % 18}",
                        "销售大区": regions[i % 6], "回款率": 0.5 + (i % 50) / 100}
                       for i in range(100)])
    wb = _load([{"sheet_name": "明细", "df": df, "chart": None}])
    assert len(wb["明细"]._charts) == 6, "100 行明细应按销售大区自动拆 6 张图"
    # 小表 → 自动单图
    wb2 = _load([{"sheet_name": "小表", "df": df.head(8)[["销售代表", "回款率"]], "chart": None}])
    assert len(wb2["小表"]._charts) == 1


def test_auto_chart_skips_pure_numeric():
    corr = pd.DataFrame({"a": [1.0, 0.5], "b": [0.5, 1.0]})
    wb = _load([{"sheet_name": "矩阵", "df": corr, "chart": None}])
    assert len(wb["矩阵"]._charts) == 0, "没有文本主体列(如相关性矩阵)不强行画图"
