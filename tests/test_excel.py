"""
Excel 输出测试 —— v4 webui writer。
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from client.webui.writer import (
    enforce_excel_path,
    make_excel_path,
    write_skill_results,
)


def test_make_excel_path_in_downloads():
    p = make_excel_path()
    assert p.name.startswith("analysis_")
    assert p.suffix == ".xlsx"
    assert "Downloads" in str(p)


def test_enforce_excel_path_accepts_downloads():
    p = Path.home() / "Downloads" / "x.xlsx"
    out = enforce_excel_path(p)
    assert out.suffix == ".xlsx"


def test_enforce_excel_path_rejects_outside(tmp_path: Path):
    bad = tmp_path / "elsewhere" / "x.xlsx"
    with pytest.raises(PermissionError):
        enforce_excel_path(bad)


def test_write_skill_results_basic(tmp_path: Path):
    df = pd.DataFrame(
        {"region": ["华东", "华南"], "amount": [100.0, 200.0]}
    )
    results = [{
        "sheet_name": "大区汇总",
        "df": df,
        "chart": {"type": "bar", "x": "region", "y": "amount", "title": "T1"},
    }]
    dst = tmp_path / "Downloads" / "x.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)
    # 用 monkeypatch 方式临时放宽路径(直接用 tmp_path 不在 ~/Downloads/ 里)
    # 简化:直接用真实 ~/Downloads/
    out = write_skill_results(results)
    assert out.exists()
    wb = load_workbook(out)
    assert "大区汇总" in wb.sheetnames
    ws = wb["大区汇总"]
    assert ws["A1"].value == "region"
    assert ws["B1"].value == "amount"
    out.unlink(missing_ok=True)


def test_write_skill_results_multi_sheet():
    results = [
        {"sheet_name": "A", "df": pd.DataFrame({"x": [1, 2]}), "chart": None},
        {"sheet_name": "B", "df": pd.DataFrame({"y": [3]}), "chart": None},
    ]
    out = write_skill_results(results)
    assert out.exists()
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"A", "B"}
    out.unlink(missing_ok=True)


def test_write_skill_results_skips_empty_df():
    results = [
        {"sheet_name": "Empty", "df": pd.DataFrame(), "chart": None},
        {"sheet_name": "OK", "df": pd.DataFrame({"x": [1]}), "chart": None},
    ]
    out = write_skill_results(results)
    wb = load_workbook(out)
    assert wb.sheetnames == ["OK"]
    out.unlink(missing_ok=True)
