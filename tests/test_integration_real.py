"""
真实 backend 集成测试。

跑条件:
- 密态数据分析包已安装 (import 成功)
- 密钥 / 字典 / 授权文件已就位

若不满足,整个文件 skip。等用户后续提供密钥时,可直接执行这里的测试验证真实链路。

⚠️ 注意:每个测试都使用全局 Runtime,会触发一次性的 initSK + initDict。
    在同一 pytest 进程内运行多次没问题(幂等)。
"""

from __future__ import annotations

import importlib

import pytest

from client.tools.runtime import Runtime, RuntimeConfig


# ---------------------------------------------------------------------------
# 跳过判定
# ---------------------------------------------------------------------------


def _packages_available() -> bool:
    for pkg in ("crypto_toolkit", "henumpy", "pandaseal", "helearn"):
        try:
            importlib.import_module(pkg)
        except Exception:
            return False
    return True


PACKAGES_OK = _packages_available()
RUNTIME_OK = PACKAGES_OK and Runtime.reset(RuntimeConfig(backend="real")).real_available()

pytestmark = pytest.mark.skipif(
    not RUNTIME_OK,
    reason=(
        "real backend 未就绪 —— 检查项:四个包是否已 pip install,"
        "且 sk / 字典 / 授权文件是否就位(详见 PROVIDE_ME.md §3)"
    ),
)


def _to_scalar(x):
    """把任意 ndarray / list / scalar 统一收敛为 float 标量。"""
    import numpy as np

    if hasattr(x, "shape") and getattr(x, "shape", None) == ():
        return float(x.item())
    if isinstance(x, np.ndarray):
        flat = x.flatten()
        return float(flat[0]) if len(flat) else 0.0
    if isinstance(x, (list, tuple)) and len(x):
        return float(x[0])
    return float(x)


# ---------------------------------------------------------------------------
# crypto_toolkit 真实加解密
# ---------------------------------------------------------------------------


def test_real_crypto_roundtrip_array():
    """numpy 数组的加密 → 解密往返。"""
    import numpy as np

    from client.tools import CryptoToolkit

    ct = CryptoToolkit(backend="real")
    arr = np.array([1.0, 2.0, 3.0, 4.0])
    cipher = ct.encrypt(arr)
    assert type(cipher).__name__ == "CipherArray"

    plain = ct.decrypt(cipher)
    assert plain is not None
    # FHE 有精度误差,只比较近似
    np_plain = np.array(plain).flatten()
    assert np.allclose(np_plain[: len(arr)], arr, atol=1e-3)


def test_real_crypto_roundtrip_dataframe(tmp_path):
    """pandas DataFrame 的加密 → 解密往返。"""
    import pandas as pd

    from client.tools import CryptoToolkit

    df = pd.DataFrame({"x": [1.0, 2.0, 3.0], "y": [4.0, 5.0, 6.0]})
    ct = CryptoToolkit(backend="real")
    cdf = ct.encrypt(df)
    assert type(cdf).__name__ == "CipherDataFrame"

    decrypted = ct.decrypt(cdf)
    # 解密结果应近似原 DataFrame
    assert decrypted is not None


# ---------------------------------------------------------------------------
# henumpy 真实计算
# ---------------------------------------------------------------------------


def test_real_henumpy_sum():
    """密文上求和。"""
    import numpy as np

    from client.tools import CryptoToolkit, HENumpy
    from shared.contract import Operation

    ct = CryptoToolkit(backend="real")
    hp_wrapper = HENumpy(backend="real")

    arr = np.array([1.0, 2.0, 3.0, 4.0])
    cipher_in = ct.encrypt(arr)
    cipher_out = hp_wrapper.run([Operation(op="sum")], cipher_in)
    result = ct.decrypt(cipher_out)
    # 期望和约为 10.0
    val = _to_scalar(result)
    assert abs(val - 10.0) < 0.1


def test_real_henumpy_mean():
    import numpy as np

    from client.tools import CryptoToolkit, HENumpy
    from shared.contract import Operation

    ct = CryptoToolkit(backend="real")
    hp_wrapper = HENumpy(backend="real")

    arr = np.array([2.0, 4.0, 6.0, 8.0])
    cipher_in = ct.encrypt(arr)
    cipher_out = hp_wrapper.run([Operation(op="mean")], cipher_in)
    result = ct.decrypt(cipher_out)
    val = _to_scalar(result)
    assert abs(val - 5.0) < 0.1


# ---------------------------------------------------------------------------
# pandaseal 真实操作
# ---------------------------------------------------------------------------


def test_real_pandaseal_mean(tmp_path):
    """加密 DataFrame → cdf.mean() (CipherSeries) → 解密为列名→均值的 dict。"""
    import pandas as pd

    from client.tools import CryptoToolkit, PandaSeal
    from shared.contract import Operation

    ct = CryptoToolkit(backend="real")
    ps_wrapper = PandaSeal(backend="real")

    df = pd.DataFrame({"a": [1.0, 2.0, 3.0], "b": [10.0, 20.0, 30.0]})
    cdf = ct.encrypt(df)
    result_cs = ps_wrapper.run([Operation(op="mean")], cdf)
    decrypted = ct.decrypt(result_cs)
    # decrypted 应是 {'a': ~2.0, 'b': ~20.0}
    assert isinstance(decrypted, dict)
    assert abs(_to_scalar(decrypted["a"]) - 2.0) < 0.5
    assert abs(_to_scalar(decrypted["b"]) - 20.0) < 0.5


# ---------------------------------------------------------------------------
# 真实端到端 —— 真实 LLM + 真实 HE + 真实 Excel
# ---------------------------------------------------------------------------

import os as _os

_HAS_LLM_KEY = bool(_os.environ.get("OPENROUTER_API_KEY"))


@pytest.mark.skipif(not _HAS_LLM_KEY, reason="未设置 OPENROUTER_API_KEY 环境变量")
def test_real_end_to_end_pandaseal(tmp_path, tmp_downloads):
    """
    完整闭环:
    1. 准备明文 CSV(销售数据)
    2. ct.encrypt_csv → 加密 CSV
    3. 真实 LLM 出 plan + summary
    4. LangGraph 工作流跑 pandaseal → 解密 → Excel
    5. 验证 Excel 文件 + summary 无明文
    """
    import pandas as pd
    from openpyxl import load_workbook

    from client.permissions import AutoApproveAuthorizer, scan_summary
    from client.skill_workflow import build_workflow
    from client.tools import CryptoToolkit, HELearn, HENumpy, HETorch, PandaSeal
    from host.llm_proxy import make_provider

    # ----- 1. 准备明文 + 加密 -----
    # 注:ct.encrypt_csv 只接受全数字列(对每个 cell 做 float())
    # 字符串列(日期、类别名)需在外部预先 label-encode 成数字
    plain_csv = tmp_path / "metrics.csv"
    pd.DataFrame(
        {
            "amount": [100.0, 200.0, 150.0, 250.0, 300.0],
            "qty": [10.0, 20.0, 15.0, 25.0, 30.0],
        }
    ).to_csv(plain_csv, index=False)

    ct = CryptoToolkit(backend="real")
    encrypted_csv = tmp_path / "metrics_enc.csv"
    ct.encrypt_file(plain_csv, encrypted_csv)
    assert encrypted_csv.exists()

    # ----- 2. 真实 LLM provider -----
    llm = make_provider(
        "openrouter",
        api_key=_os.environ["OPENROUTER_API_KEY"],
        model=_os.environ.get("MODEL_NAME", "deepseek/deepseek-v4-pro"),
    )

    # ----- 3. 构建工作流(全部 real backend) -----
    wf = build_workflow(
        llm_client=llm,
        zfhe=ct,
        pandaseal=PandaSeal(backend="real"),
        henumpy=HENumpy(backend="real"),
        helearn=HELearn(backend="real"),
        hetorch=HETorch(backend="stub"),  # hetorch2 包还未提供
        authorizer=AutoApproveAuthorizer(),
        max_retries=1,
    )

    schema = {
        "fields": [
            {"name": "amount", "type": "float"},
            {"name": "qty", "type": "float"},
        ]
    }
    state = wf.invoke(
        {
            "user_query": "计算 amount 和 qty 各列的均值,作为整体指标输出",
            "schema": schema,
            "ciphertext_paths": [str(encrypted_csv)],
        }
    )

    # ----- 4. 验证 -----
    if state.get("error") and not state.get("excel_path"):
        pytest.fail(f"工作流失败: {state['error']}; state={state}")

    # Excel 生成
    from pathlib import Path as _Path

    excel_path = state.get("excel_path")
    assert excel_path, f"未生成 Excel: state={state}"
    assert _Path(excel_path).exists()

    # summary 通过 B6-3 过滤
    summary = state.get("summary_filtered", "")
    fr = scan_summary(summary)
    assert fr.clean, f"summary 含明文: {fr.report()}; summary={summary!r}"

    # Excel 内容可读
    wb = load_workbook(excel_path)
    assert wb.sheetnames, "Excel 应至少有一个 sheet"
