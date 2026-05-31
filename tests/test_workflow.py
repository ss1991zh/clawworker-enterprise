"""
LangGraph 端到端工作流测试。

覆盖:
- 场景 1 描述性分析,成功路径
- 场景 2 数值计算,成功路径
- summary 命中红线 → 重试 → 仍命中 → fallback
- 解密授权被拒 → 终止
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from client.permissions import AutoApproveAuthorizer, DenyAuthorizer
from client.skill_workflow import build_workflow
from client.tools import HELearn, HENumpy, HETorch, PandaSeal, ZFHE
from shared.contract import (
    ChartSpec,
    ComputationPlan,
    ExcelOutput,
    LLMResponse,
    Operation,
    Scenario,
    SheetSpec,
)


# ---------------------------------------------------------------------------
# 场景 1:描述性分析端到端
# ---------------------------------------------------------------------------


def test_scenario_1_end_to_end(
    tmp_downloads: Path,
    zfhe: ZFHE,
    pandaseal: PandaSeal,
    henumpy: HENumpy,
    helearn: HELearn,
    hetorch: HETorch,
    sample_ciphertext_file: Path,
    fixed_llm_factory,
):
    plan = ComputationPlan(
        scenario=Scenario.DESCRIPTIVE,
        tool="pandaseal",
        ops=[
            Operation(op="group_by", field="month"),
            Operation(op="sum", field="amount"),
        ],
        output=ExcelOutput(
            file="~/Downloads/analysis_test.xlsx",  # 字面路径,实际写入走 make_excel_path
            sheets=[
                SheetSpec(
                    name="MonthlyTrend",
                    columns=["group", "amount_sum"],
                    chart=ChartSpec(type="line", x="group", y="amount_sum"),
                )
            ],
        ),
    )
    llm_resp = LLMResponse(
        computation_plan=plan,
        summary="已按月份聚合销售额并生成折线图,详见 Excel 的 MonthlyTrend sheet。",
    )
    llm = fixed_llm_factory([llm_resp])

    wf = build_workflow(
        llm_client=llm,
        zfhe=zfhe,
        pandaseal=pandaseal,
        henumpy=henumpy,
        helearn=helearn,
        hetorch=hetorch,
        authorizer=AutoApproveAuthorizer(),
        system_prompt="<test system prompt>",
    )

    state = wf.invoke(
        {
            "user_query": "按月份汇总销售额",
            "schema": {"fields": [{"name": "month", "type": "string"}, {"name": "amount", "type": "int"}]},
            "ciphertext_paths": [str(sample_ciphertext_file)],
        }
    )

    # 验证最终状态
    assert state.get("excel_path"), f"未生成 Excel: state={state}"
    assert state.get("summary_filtered")
    assert state.get("summary_filter_hit") is False
    excel_path = Path(state["excel_path"])
    assert excel_path.exists()

    # 验证 Excel 内容
    wb = load_workbook(excel_path)
    ws = wb["MonthlyTrend"]
    assert ws["A1"].value == "group"
    assert ws["B1"].value == "amount_sum"
    # 数据行
    rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)]
    by_month = dict(rows)
    assert by_month["2024-01"] == 300
    assert by_month["2024-02"] == 400
    assert by_month["2024-03"] == 300


# ---------------------------------------------------------------------------
# 场景 2:数值计算
# ---------------------------------------------------------------------------


def test_scenario_2_correlation(
    tmp_downloads: Path,
    zfhe: ZFHE,
    pandaseal: PandaSeal,
    henumpy: HENumpy,
    helearn: HELearn,
    hetorch: HETorch,
    tmp_path: Path,
    fixed_llm_factory,
):
    # 准备密文数据
    cipher = zfhe.encrypt({"x": [1, 2, 3, 4], "y": [2, 4, 6, 8]})
    data_path = tmp_path / "matrix.cipher"
    data_path.write_bytes(cipher)

    plan = ComputationPlan(
        scenario=Scenario.NUMERICAL,
        tool="henumpy",
        ops=[Operation(op="corrcoef")],
        output=ExcelOutput(
            file="~/Downloads/numeric.xlsx",
            sheets=[SheetSpec(name="Corr")],
        ),
    )
    llm_resp = LLMResponse(
        computation_plan=plan,
        summary="已计算字段间相关系数矩阵,详见 Excel 的 Corr sheet。",
    )

    wf = build_workflow(
        llm_client=fixed_llm_factory([llm_resp]),
        zfhe=zfhe,
        pandaseal=pandaseal,
        henumpy=henumpy,
        helearn=helearn,
        hetorch=hetorch,
        authorizer=AutoApproveAuthorizer(),
        system_prompt="<test>",
    )
    state = wf.invoke(
        {
            "user_query": "x 和 y 相关性",
            "schema": {"fields": [{"name": "x", "type": "float"}, {"name": "y", "type": "float"}]},
            "ciphertext_paths": [str(data_path)],
        }
    )
    assert state.get("excel_path")
    wb = load_workbook(state["excel_path"])
    ws = wb["Corr"]
    assert ws["A1"].value == "label"


# ---------------------------------------------------------------------------
# summary 命中红线 → 限次重试 → fallback
# ---------------------------------------------------------------------------


def test_summary_filter_triggers_retry_then_fallback(
    tmp_downloads: Path,
    zfhe: ZFHE,
    pandaseal: PandaSeal,
    henumpy: HENumpy,
    helearn: HELearn,
    hetorch: HETorch,
    sample_ciphertext_file: Path,
    fixed_llm_factory,
):
    plan = ComputationPlan(
        scenario=Scenario.DESCRIPTIVE,
        tool="pandaseal",
        ops=[Operation(op="group_by", field="month"), Operation(op="sum", field="amount")],
        output=ExcelOutput(
            file="~/Downloads/x.xlsx",
            sheets=[SheetSpec(name="X", columns=["group", "amount_sum"])],
        ),
    )
    # 三次响应都"忍不住举例" —— 都应该被拦截,最终走 fallback
    bad_summary = "2024 年 11 月销售额达 120 万,增长 30%。"
    responses = [
        LLMResponse(computation_plan=plan, summary=bad_summary),
        LLMResponse(computation_plan=plan, summary=bad_summary),
        LLMResponse(computation_plan=plan, summary=bad_summary),
    ]
    llm = fixed_llm_factory(responses)

    wf = build_workflow(
        llm_client=llm,
        zfhe=zfhe,
        pandaseal=pandaseal,
        henumpy=henumpy,
        helearn=helearn,
        hetorch=hetorch,
        authorizer=AutoApproveAuthorizer(),
        max_retries=2,
        system_prompt="<test>",
    )
    state = wf.invoke(
        {
            "user_query": "按月份汇总",
            "schema": {"fields": []},
            "ciphertext_paths": [str(sample_ciphertext_file)],
        }
    )

    # 应该用了 fallback,且 LLM 被调用了 max_retries+1 次
    assert len(llm.calls) == 3, f"LLM 调用次数不对: {len(llm.calls)}"
    assert state.get("summary_filtered") == (
        "分析已完成,具体数据与图表请打开本次生成的 Excel 文件查看。"
        "(为保护数据隐私,聊天界面不会显示任何具体内容)"
    )
    # Excel 仍应生成(fallback 路径走完全部计算)
    assert state.get("excel_path")


# ---------------------------------------------------------------------------
# 解密授权被拒
# ---------------------------------------------------------------------------


def test_authorization_denied_produces_encrypted_excel(
    tmp_downloads: Path,
    zfhe: ZFHE,
    pandaseal: PandaSeal,
    henumpy: HENumpy,
    helearn: HELearn,
    hetorch: HETorch,
    sample_ciphertext_file: Path,
    fixed_llm_factory,
):
    """新行为:授权被拒 → 仍产出 Excel,但内容为序列化后的密文。"""
    plan = ComputationPlan(
        scenario=Scenario.DESCRIPTIVE,
        tool="pandaseal",
        ops=[Operation(op="group_by", field="month"), Operation(op="sum", field="amount")],
        output=ExcelOutput(
            file="~/Downloads/x.xlsx",
            sheets=[SheetSpec(name="X")],
        ),
    )
    llm_resp = LLMResponse(computation_plan=plan, summary="已完成分析,详见 Excel。")
    wf = build_workflow(
        llm_client=fixed_llm_factory([llm_resp]),
        zfhe=zfhe,
        pandaseal=pandaseal,
        henumpy=henumpy,
        helearn=helearn,
        hetorch=hetorch,
        authorizer=DenyAuthorizer(),
        system_prompt="<test>",
    )
    state = wf.invoke(
        {
            "user_query": "...",
            "schema": {"fields": []},
            "ciphertext_paths": [str(sample_ciphertext_file)],
        }
    )

    # 授权被拒
    assert state.get("authorized") is False
    # 但 Excel 仍产出
    assert state.get("excel_path"), f"未产出密文 Excel: {state}"
    assert Path(state["excel_path"]).exists()
    # 文件名应带 _encrypted 前缀
    assert "encrypted" in Path(state["excel_path"]).name
    # summary 提示数据是密文
    assert "密文" in (state.get("summary_filtered") or "")

    # Excel 里应含 ciphertext 列(stub backend 写的是 bytes 形式)
    from openpyxl import load_workbook

    wb = load_workbook(state["excel_path"])
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "ciphertext" in headers, f"Excel 无 ciphertext 列: {headers}"
